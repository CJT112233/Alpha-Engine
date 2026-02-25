import io
import re
import logging
from datetime import datetime
from typing import Optional

logger = logging.getLogger(__name__)

TYPE_LABELS = {
    "A": "Wastewater Treatment",
    "B": "RNG Greenfield",
    "C": "RNG Bolt-On",
    "D": "Hybrid",
}


def _sanitize(text) -> str:
    if not text:
        return ""
    s = str(text)
    s = s.replace("\u2018", "'").replace("\u2019", "'").replace("\u201a", "'")
    s = s.replace("\u201c", '"').replace("\u201d", '"').replace("\u201e", '"')
    s = s.replace("\u2026", "...").replace("\u2013", "-").replace("\u2014", "--")
    s = s.replace("\u00a0", " ")
    return s


def _fmt_num(val, decimals=1) -> str:
    if val is None:
        return "-"
    try:
        v = float(val)
        return f"{v:,.{decimals}f}"
    except (ValueError, TypeError):
        return str(val)


def _fmt_currency(val) -> str:
    try:
        v = float(val)
        return f"${v:,.0f}"
    except (ValueError, TypeError):
        return str(val)


def _camel_to_title(s: str) -> str:
    result = re.sub(r"([A-Z])", r" \1", s)
    return result.strip().title()


def _wrap_text(text: str, max_width: float, font_name: str, font_size: int) -> list:
    from reportlab.pdfbase.pdfmetrics import stringWidth
    words = text.split()
    lines = []
    current = ""
    for word in words:
        test = f"{current} {word}".strip()
        if stringWidth(test, font_name, font_size) <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines if lines else [""]


def _draw_table(c, headers, rows, x, y, col_widths,
                font_size=8, header_bg="#2563EB", page_height=792, page_width=612):
    from reportlab.lib.colors import HexColor
    min_row_h = 16
    pad = 3
    table_w = sum(col_widths)

    def measure_h(cells, bold):
        from reportlab.pdfbase.pdfmetrics import stringWidth
        max_h = min_row_h
        for i, cell in enumerate(cells):
            cw = col_widths[i] - pad * 2
            fn = "Helvetica-Bold" if bold else "Helvetica"
            words = _sanitize(cell or "").split()
            lines = 1
            current = ""
            for w in words:
                test = f"{current} {w}".strip()
                if stringWidth(test, fn, font_size) > cw:
                    lines += 1
                    current = w
                else:
                    current = test
            h = (lines * (font_size + 2)) + pad * 2
            if h > max_h:
                max_h = h
        return max_h

    def draw_row(cells, bold, bg=None):
        nonlocal y
        rh = measure_h(cells, bold)
        if y - rh < 60:
            c.showPage()
            y = page_height - 50
            draw_row(headers, True, header_bg)
            return
        if bg:
            c.setFillColor(HexColor(bg))
            c.rect(x, y - rh, table_w, rh, fill=1, stroke=0)
        cx = x
        for i, cell in enumerate(cells):
            fn = "Helvetica-Bold" if bold else "Helvetica"
            fc = "#FFFFFF" if bg == header_bg else "#333333"
            c.setFont(fn, font_size)
            c.setFillColor(HexColor(fc))
            text = _sanitize(cell or "")
            cw = col_widths[i] - pad * 2
            wrapped = _wrap_text(text, cw, fn, font_size)
            ty = y - pad - font_size
            for line in wrapped:
                if ty < y - rh + pad:
                    break
                c.drawString(cx + pad, ty, line)
                ty -= (font_size + 2)
            cx += col_widths[i]
        c.setStrokeColor(HexColor("#CCCCCC"))
        c.setLineWidth(0.5)
        c.rect(x, y - rh, table_w, rh, fill=0, stroke=1)
        y -= rh

    draw_row(headers, True, header_bg)
    for idx, row in enumerate(rows):
        safe_row = [str(cell) if cell is not None else "-" for cell in row]
        bg = "#F8F9FA" if idx % 2 == 1 else None
        draw_row(safe_row, False, bg)
    return y


def _add_section_header(c, title, y, left_margin, content_width, page_height=792):
    from reportlab.lib.colors import HexColor
    if y < 100:
        c.showPage()
        y = page_height - 50
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(HexColor("#1E3A5F"))
    c.drawString(left_margin, y, title)
    y -= 20
    c.setStrokeColor(HexColor("#CCCCCC"))
    c.setLineWidth(0.5)
    c.line(left_margin, y, left_margin + content_width, y)
    y -= 8
    return y


def export_mass_balance_pdf(results: dict, scenario_name: str, project_name: str, project_type: str) -> bytes:
    from reportlab.lib.pagesizes import LETTER
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import HexColor

    page_width, page_height = LETTER
    left_margin = 50
    content_width = page_width - 100

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=LETTER)

    c.setFont("Helvetica-Bold", 18)
    c.setFillColor(HexColor("#1E3A5F"))
    c.drawCentredString(page_width / 2, page_height - 50, "Mass Balance Report")

    c.setFont("Helvetica", 10)
    c.setFillColor(HexColor("#666666"))
    c.drawCentredString(page_width / 2, page_height - 75, f"Project: {_sanitize(project_name)}")
    c.drawCentredString(page_width / 2, page_height - 88, f"Scenario: {_sanitize(scenario_name)}")
    c.drawCentredString(page_width / 2, page_height - 101, f"Type {project_type}: {TYPE_LABELS.get(project_type, project_type)}")
    c.drawCentredString(page_width / 2, page_height - 114, f"Generated: {datetime.now().strftime('%m/%d/%Y')}")

    y = page_height - 140

    summary = results.get("summary") or {}
    if isinstance(summary, dict) and len(summary) > 0:
        y = _add_section_header(c, "Summary", y, left_margin, content_width, page_height)
        headers = ["Parameter", "Value", "Unit"]
        rows = []
        for key, val in summary.items():
            if isinstance(val, dict):
                rows.append([_camel_to_title(key), str(val.get("value", "-")), str(val.get("unit", ""))])
            else:
                rows.append([_camel_to_title(key), str(val), ""])
        y = _draw_table(c, headers, rows, left_margin, y, [200, 162, 150], page_height=page_height, page_width=page_width)
        y -= 15

    stages = results.get("stages") or []
    if isinstance(stages, list) and len(stages) > 0:
        y = _add_section_header(c, "Treatment Train - Stream Data", y, left_margin, content_width, page_height)
        conv = results.get("convergenceAchieved")
        if conv is not None:
            c.setFont("Helvetica", 8)
            c.setFillColor(HexColor("#666666"))
            c.drawString(left_margin, y, f"Convergence: {'Yes' if conv else 'No'} ({results.get('convergenceIterations', 0)} iterations)")
            y -= 14

        param_labels = {
            "flow": "Flow (GPD)", "bod": "BOD (mg/L)", "cod": "COD (mg/L)",
            "tss": "TSS (mg/L)", "tkn": "TKN (mg/L)", "tp": "TP (mg/L)", "fog": "FOG (mg/L)",
        }
        params = ["flow", "bod", "cod", "tss", "tkn", "tp", "fog"]

        for stage in stages:
            if y < 120:
                c.showPage()
                y = page_height - 50
            c.setFont("Helvetica-Bold", 9)
            c.setFillColor(HexColor("#333333"))
            c.drawString(left_margin, y, f"{_sanitize(stage.get('name', ''))} ({_sanitize(stage.get('type', ''))})")
            y -= 14

            s_headers = ["Parameter", "Influent", "Effluent", "Removal %"]
            influent = stage.get("influent") or {}
            effluent = stage.get("effluent") or {}
            removals = stage.get("removalEfficiencies") or {}
            s_rows = []
            for p in params:
                s_rows.append([
                    param_labels.get(p, p.upper()),
                    _fmt_num(influent.get(p)),
                    _fmt_num(effluent.get(p)),
                    f"{_fmt_num(removals.get(p))}%" if removals.get(p) is not None else "-",
                ])
            y = _draw_table(c, s_headers, s_rows, left_margin, y, [128, 128, 128, 128], page_height=page_height, page_width=page_width)
            y -= 10

    ad_stages = results.get("adStages") or []
    if isinstance(ad_stages, list) and len(ad_stages) > 0:
        y = _add_section_header(c, "AD / RNG Process Stages", y, left_margin, content_width, page_height)
        for stage in ad_stages:
            if y < 120:
                c.showPage()
                y = page_height - 50
            c.setFont("Helvetica-Bold", 9)
            c.setFillColor(HexColor("#333333"))
            c.drawString(left_margin, y, f"{_sanitize(stage.get('name', ''))} ({_sanitize(stage.get('type', ''))})")
            y -= 14

            input_stream = stage.get("inputStream") or {}
            output_stream = stage.get("outputStream") or {}
            all_keys = list(dict.fromkeys(list(input_stream.keys()) + list(output_stream.keys())))

            ad_headers = ["Parameter", "Input", "Output"]
            ad_rows = []
            for key in all_keys:
                inp = input_stream.get(key)
                out = output_stream.get(key)
                ad_rows.append([
                    _camel_to_title(key),
                    f"{_fmt_num(inp.get('value'))} {inp.get('unit', '')}" if isinstance(inp, dict) else "-",
                    f"{_fmt_num(out.get('value'))} {out.get('unit', '')}" if isinstance(out, dict) else "-",
                ])
            y = _draw_table(c, ad_headers, ad_rows, left_margin, y, [171, 171, 170], page_height=page_height, page_width=page_width)
            y -= 10

    equipment = results.get("equipment") or []
    if isinstance(equipment, list) and len(equipment) > 0:
        y = _add_section_header(c, "Equipment List", y, left_margin, content_width, page_height)
        eq_headers = ["Process", "Equipment", "Qty", "Description", "Design Basis"]
        eq_rows = [[
            _sanitize(eq.get("process", "")),
            _sanitize(eq.get("equipmentType", "")),
            str(eq.get("quantity", 1)),
            _sanitize(eq.get("description", "")),
            _sanitize(eq.get("designBasis", "")),
        ] for eq in equipment]
        y = _draw_table(c, eq_headers, eq_rows, left_margin, y, [90, 100, 32, 160, 130], page_height=page_height, page_width=page_width)
        y -= 15

    assumptions = results.get("assumptions") or []
    if isinstance(assumptions, list) and len(assumptions) > 0:
        y = _add_section_header(c, "Design Assumptions", y, left_margin, content_width, page_height)
        ass_headers = ["Parameter", "Value", "Source"]
        ass_rows = [[_sanitize(a.get("parameter", "")), _sanitize(a.get("value", "")), _sanitize(a.get("source", ""))] for a in assumptions]
        y = _draw_table(c, ass_headers, ass_rows, left_margin, y, [180, 162, 170], page_height=page_height, page_width=page_width)

    c.save()
    buf.seek(0)
    return buf.read()


def export_mass_balance_excel(results: dict, scenario_name: str, project_name: str, project_type: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    summary = results.get("summary") or {}
    if isinstance(summary, dict) and len(summary) > 0:
        ws = wb.create_sheet("Summary")
        ws.append(["Mass Balance Summary"])
        ws.append([f"Project: {project_name}", f"Scenario: {scenario_name}", f"Type: {project_type} - {TYPE_LABELS.get(project_type, '')}"])
        ws.append([])
        ws.append(["Parameter", "Value", "Unit"])
        for key, val in summary.items():
            if isinstance(val, dict):
                ws.append([_camel_to_title(key), val.get("value", "-"), val.get("unit", "")])
            else:
                ws.append([_camel_to_title(key), val, ""])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20

    stages = results.get("stages") or []
    if isinstance(stages, list) and len(stages) > 0:
        ws = wb.create_sheet("Treatment Train")
        ws.append(["Treatment Train - Stream Data"])
        conv = results.get("convergenceAchieved")
        ws.append([f"Convergence: {'Yes' if conv else 'No'} ({results.get('convergenceIterations', 0)} iterations)"])
        ws.append([])

        param_labels = {
            "flow": "Flow (GPD)", "bod": "BOD (mg/L)", "cod": "COD (mg/L)",
            "tss": "TSS (mg/L)", "tkn": "TKN (mg/L)", "tp": "TP (mg/L)", "fog": "FOG (mg/L)",
        }
        params = ["flow", "bod", "cod", "tss", "tkn", "tp", "fog"]

        for stage in stages:
            ws.append([f"{stage.get('name', '')} ({stage.get('type', '')})"])
            ws.append(["Parameter", "Influent", "Effluent", "Removal %"])
            influent = stage.get("influent") or {}
            effluent = stage.get("effluent") or {}
            removals = stage.get("removalEfficiencies") or {}
            for p in params:
                ws.append([
                    param_labels.get(p, p.upper()),
                    influent.get(p, ""),
                    effluent.get(p, ""),
                    removals.get(p, ""),
                ])
            ws.append([])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 15

    ad_stages = results.get("adStages") or []
    if isinstance(ad_stages, list) and len(ad_stages) > 0:
        ws = wb.create_sheet("AD Process")
        ws.append(["AD / RNG Process Stages"])
        ws.append([])
        for stage in ad_stages:
            ws.append([f"{stage.get('name', '')} ({stage.get('type', '')})"])
            ws.append(["Parameter", "Input Value", "Input Unit", "Output Value", "Output Unit"])
            input_stream = stage.get("inputStream") or {}
            output_stream = stage.get("outputStream") or {}
            all_keys = list(dict.fromkeys(list(input_stream.keys()) + list(output_stream.keys())))
            for key in all_keys:
                inp = input_stream.get(key)
                out = output_stream.get(key)
                ws.append([
                    _camel_to_title(key),
                    inp.get("value", "") if isinstance(inp, dict) else "",
                    inp.get("unit", "") if isinstance(inp, dict) else "",
                    out.get("value", "") if isinstance(out, dict) else "",
                    out.get("unit", "") if isinstance(out, dict) else "",
                ])
            ws.append([])
        for col_letter, w in [("A", 25), ("B", 15), ("C", 15), ("D", 15), ("E", 15)]:
            ws.column_dimensions[col_letter].width = w

    equipment = results.get("equipment") or []
    if isinstance(equipment, list) and len(equipment) > 0:
        ws = wb.create_sheet("Equipment")
        ws.append(["Equipment List"])
        ws.append([])
        ws.append(["Process", "Equipment Type", "Qty", "Description", "Design Basis", "Notes"])
        for eq in equipment:
            ws.append([
                eq.get("process", ""),
                eq.get("equipmentType", ""),
                eq.get("quantity", 1),
                eq.get("description", ""),
                eq.get("designBasis", ""),
                eq.get("notes", ""),
            ])
        for col_letter, w in [("A", 20), ("B", 20), ("C", 6), ("D", 35), ("E", 30), ("F", 30)]:
            ws.column_dimensions[col_letter].width = w

    assumptions = results.get("assumptions") or []
    if isinstance(assumptions, list) and len(assumptions) > 0:
        ws = wb.create_sheet("Assumptions")
        ws.append(["Design Assumptions"])
        ws.append([])
        ws.append(["Parameter", "Value", "Source"])
        for a in assumptions:
            ws.append([a.get("parameter", ""), a.get("value", ""), a.get("source", "")])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def export_capex_pdf(results: dict, scenario_name: str, project_name: str, project_type: str) -> bytes:
    from reportlab.lib.pagesizes import LETTER
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import HexColor

    page_width, page_height = LETTER
    left_margin = 50
    content_width = page_width - 100

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=LETTER)

    c.setFont("Helvetica-Bold", 18)
    c.setFillColor(HexColor("#1E3A5F"))
    c.drawCentredString(page_width / 2, page_height - 50, "Capital Cost Estimate")

    c.setFont("Helvetica", 10)
    c.setFillColor(HexColor("#666666"))
    c.drawCentredString(page_width / 2, page_height - 75, f"Project: {_sanitize(project_name)}")
    c.drawCentredString(page_width / 2, page_height - 88, f"Scenario: {_sanitize(scenario_name)}")
    c.drawCentredString(page_width / 2, page_height - 101, f"Type {project_type}: {TYPE_LABELS.get(project_type, project_type)}")
    cost_year = results.get("costYear", "Current")
    currency = results.get("currency", "USD")
    c.drawCentredString(page_width / 2, page_height - 114, f"Cost Year: {cost_year} | Currency: {currency}")
    c.drawCentredString(page_width / 2, page_height - 127, f"Generated: {datetime.now().strftime('%m/%d/%Y')}")

    y = page_height - 155

    summary = results.get("summary") or {}
    if isinstance(summary, dict) and len(summary) > 0:
        y = _add_section_header(c, "Cost Summary", y, left_margin, content_width, page_height)
        equipment_cost = summary.get("totalEquipmentCost", 0)
        is_deterministic = "subtotalDirectCosts" in summary
        total_direct = summary.get("subtotalDirectCosts", 0) if is_deterministic else summary.get("totalInstalledCost", 0)
        installation_cost = total_direct - equipment_cost
        eng_cost = summary.get("engineeringCost", 0)
        contingency_amt = summary.get("contingency", summary.get("totalContingency", 0))
        total_capital = summary.get("totalProjectCost", 0)
        other_indirect = total_capital - total_direct - eng_cost - contingency_amt
        sum_rows = [
            ["Equipment", _fmt_currency(equipment_cost)],
            ["Installation", _fmt_currency(installation_cost)],
            ["Total Direct Costs", _fmt_currency(total_direct)],
            [f"Engineering ({summary.get('engineeringPct', 7)}%)", _fmt_currency(eng_cost)],
            ["Other Indirect Costs", _fmt_currency(max(other_indirect, 0))],
            ["Contingency (7.5%)", _fmt_currency(contingency_amt)],
            ["Total Capital Costs", _fmt_currency(total_capital)],
        ]
        cpu = summary.get("costPerUnit")
        if isinstance(cpu, dict):
            sum_rows.append([f"Cost per Unit ({cpu.get('basis', '')})", f"{_fmt_currency(cpu.get('value', 0))} / {cpu.get('unit', '')}"])
        y = _draw_table(c, ["Item", "Amount"], sum_rows, left_margin, y, [300, 212],
                        header_bg="#1E3A5F", page_height=page_height, page_width=page_width)
        y -= 15

    line_items = results.get("lineItems") or []
    if isinstance(line_items, list) and len(line_items) > 0:
        y = _add_section_header(c, "Line Items", y, left_margin, content_width, page_height)
        li_headers = ["Process", "Equipment", "Qty", "Base Cost", "Install Factor", "Installed", "Contingency", "Total"]
        li_rows = [[
            _sanitize(li.get("process", "")),
            _sanitize(li.get("equipmentType", "")),
            str(li.get("quantity", 1)),
            _fmt_currency(li.get("baseCostPerUnit", 0)),
            f"{_fmt_num(li.get('installationFactor', 0), 2)}x",
            _fmt_currency(li.get("installedCost", 0)),
            _fmt_currency(li.get("contingencyCost", 0)),
            _fmt_currency(li.get("totalCost", 0)),
        ] for li in line_items]
        y = _draw_table(c, li_headers, li_rows, left_margin, y, [80, 80, 28, 64, 52, 72, 64, 72],
                        font_size=7, page_height=page_height, page_width=page_width)
        y -= 15

    assumptions = results.get("assumptions") or []
    if isinstance(assumptions, list) and len(assumptions) > 0:
        y = _add_section_header(c, "Assumptions", y, left_margin, content_width, page_height)
        ass_headers = ["Parameter", "Value", "Source"]
        ass_rows = [[_sanitize(a.get("parameter", "")), _sanitize(a.get("value", "")), _sanitize(a.get("source", ""))] for a in assumptions]
        y = _draw_table(c, ass_headers, ass_rows, left_margin, y, [180, 162, 170], page_height=page_height, page_width=page_width)

    methodology = results.get("methodology")
    if methodology:
        if y < 100:
            c.showPage()
            y = page_height - 50
        y = _add_section_header(c, "Methodology", y, left_margin, content_width, page_height)
        c.setFont("Helvetica", 9)
        c.setFillColor(HexColor("#333333"))
        for line in _wrap_text(_sanitize(methodology), content_width, "Helvetica", 9):
            c.drawString(left_margin, y, line)
            y -= 12

    c.save()
    buf.seek(0)
    return buf.read()


def export_capex_excel(results: dict, scenario_name: str, project_name: str, project_type: str) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    summary = results.get("summary") or {}
    if isinstance(summary, dict) and len(summary) > 0:
        ws = wb.create_sheet("Summary")
        ws.append(["Capital Cost Estimate - Summary"])
        ws.append([f"Project: {project_name}", f"Scenario: {scenario_name}", f"Type: {project_type} - {TYPE_LABELS.get(project_type, '')}"])
        cost_year = results.get("costYear", "Current")
        currency = results.get("currency", "USD")
        ws.append([f"Cost Year: {cost_year}", f"Currency: {currency}"])
        ws.append([])
        ws.append(["Item", "Amount ($)"])
        eq_cost = summary.get("totalEquipmentCost", 0)
        is_det = "subtotalDirectCosts" in summary
        td_cost = summary.get("subtotalDirectCosts", 0) if is_det else summary.get("totalInstalledCost", 0)
        inst_cost = td_cost - eq_cost
        eng_c = summary.get("engineeringCost", 0)
        cont_c = summary.get("contingency", summary.get("totalContingency", 0))
        tc_cost = summary.get("totalProjectCost", 0)
        oi_cost = tc_cost - td_cost - eng_c - cont_c
        ws.append(["Equipment", eq_cost])
        ws.append(["Installation", inst_cost])
        ws.append(["Total Direct Costs", td_cost])
        ws.append([f"Engineering ({summary.get('engineeringPct', 7)}%)", eng_c])
        ws.append(["Other Indirect Costs", max(oi_cost, 0)])
        ws.append(["Contingency (7.5%)", cont_c])
        ws.append(["Total Capital Costs", tc_cost])
        cpu = summary.get("costPerUnit")
        if isinstance(cpu, dict):
            ws.append([f"Cost per Unit ({cpu.get('basis', '')} - {cpu.get('unit', '')})", cpu.get("value", 0)])
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 30

    line_items = results.get("lineItems") or []
    if isinstance(line_items, list) and len(line_items) > 0:
        ws = wb.create_sheet("Line Items")
        ws.append(["Line Items"])
        ws.append([])
        ws.append([
            "Process", "Equipment Type", "Description", "Qty",
            "Base Cost/Unit ($)", "Installation Factor", "Installed Cost ($)",
            "Contingency %", "Contingency ($)", "Total Cost ($)",
            "Cost Basis", "Source", "Notes",
        ])
        for li in line_items:
            ws.append([
                li.get("process", ""),
                li.get("equipmentType", ""),
                li.get("description", ""),
                li.get("quantity", 1),
                li.get("baseCostPerUnit", 0),
                li.get("installationFactor", 0),
                li.get("installedCost", 0),
                li.get("contingencyPct", 0),
                li.get("contingencyCost", 0),
                li.get("totalCost", 0),
                li.get("costBasis", ""),
                li.get("source", ""),
                li.get("notes", ""),
            ])
        for col_letter, w in [("A", 18), ("B", 20), ("C", 30), ("D", 6),
                               ("E", 15), ("F", 12), ("G", 15), ("H", 10),
                               ("I", 15), ("J", 15), ("K", 20), ("L", 20), ("M", 25)]:
            ws.column_dimensions[col_letter].width = w

    assumptions = results.get("assumptions") or []
    if isinstance(assumptions, list) and len(assumptions) > 0:
        ws = wb.create_sheet("Assumptions")
        ws.append(["Assumptions"])
        ws.append([])
        ws.append(["Parameter", "Value", "Source"])
        for a in assumptions:
            ws.append([a.get("parameter", ""), a.get("value", ""), a.get("source", "")])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def generate_project_summary_pdf(
    project_name: str,
    scenario_name: str,
    project_type: str,
    upif: dict = None,
    mb_results: dict = None,
    capex_results: dict = None,
    opex_results: dict = None,
    financial_results: dict = None,
) -> bytes:
    from reportlab.lib.pagesizes import LETTER
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import HexColor

    page_width, page_height = LETTER
    left_margin = 50
    content_width = page_width - 100

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=LETTER)

    c.setFont("Helvetica-Bold", 22)
    c.setFillColor(HexColor("#1E3A5F"))
    c.drawCentredString(page_width / 2, page_height - 50, "PROJECT SUMMARY")

    c.setFont("Helvetica", 11)
    c.setFillColor(HexColor("#666666"))
    c.drawCentredString(page_width / 2, page_height - 78, f"Project: {_sanitize(project_name)}")
    c.drawCentredString(page_width / 2, page_height - 93, f"Scenario: {_sanitize(scenario_name)}")
    type_label = TYPE_LABELS.get(project_type, project_type)
    c.drawCentredString(page_width / 2, page_height - 108, f"Type {project_type}: {type_label}")
    c.drawCentredString(page_width / 2, page_height - 123, f"Generated: {datetime.now().strftime('%m/%d/%Y')}")

    c.setStrokeColor(HexColor("#2563EB"))
    c.setLineWidth(2)
    c.line(left_margin, page_height - 135, left_margin + content_width, page_height - 135)

    y = page_height - 160

    if upif and isinstance(upif, dict):
        y = _add_section_header(c, "Project Overview", y, left_margin, content_width, page_height)
        overview_rows = []
        feedstock_type = upif.get("feedstockType") or upif.get("feedstock_type") or ""
        if feedstock_type:
            overview_rows.append(["Feedstock Type", _sanitize(str(feedstock_type))])
        volume = upif.get("volume") or upif.get("feedstockVolume") or ""
        unit = upif.get("unit") or upif.get("feedstockUnit") or ""
        if volume:
            vol_str = f"{_fmt_num(volume)} {_sanitize(str(unit))}" if unit else _fmt_num(volume)
            overview_rows.append(["Volume", vol_str])
        location = upif.get("location") or ""
        if location:
            overview_rows.append(["Location", _sanitize(str(location))])
        output_req = upif.get("outputRequirements") or upif.get("output_requirements") or ""
        if output_req:
            if isinstance(output_req, list):
                overview_rows.append(["Output Requirements", ", ".join([_sanitize(str(r)) for r in output_req])])
            else:
                overview_rows.append(["Output Requirements", _sanitize(str(output_req))])

        if overview_rows:
            y = _draw_table(c, ["Parameter", "Value"], overview_rows, left_margin, y,
                            [200, 312], header_bg="#2563EB", page_height=page_height, page_width=page_width)
            y -= 10

        feedstocks = upif.get("feedstocks") or []
        if isinstance(feedstocks, list) and len(feedstocks) > 0:
            if y < 100:
                c.showPage()
                y = page_height - 50
            c.setFont("Helvetica-Bold", 10)
            c.setFillColor(HexColor("#333333"))
            c.drawString(left_margin, y, "Feedstock Breakdown:")
            y -= 14
            fs_rows = []
            for fs in feedstocks:
                if isinstance(fs, dict):
                    fs_name = _sanitize(str(fs.get("name", fs.get("type", ""))))
                    fs_vol = _fmt_num(fs.get("volume", fs.get("quantity", "")))
                    fs_rows.append([fs_name, fs_vol])
                else:
                    fs_rows.append([_sanitize(str(fs)), ""])
            y = _draw_table(c, ["Feedstock", "Volume"], fs_rows, left_margin, y,
                            [300, 212], header_bg="#2563EB", page_height=page_height, page_width=page_width)
            y -= 10
        y -= 5

    if mb_results and isinstance(mb_results, dict):
        y = _add_section_header(c, "Mass Balance Summary", y, left_margin, content_width, page_height)
        summary = mb_results.get("summary") or {}
        if isinstance(summary, dict) and len(summary) > 0:
            mb_rows = []
            for key, val in summary.items():
                if isinstance(val, dict):
                    mb_rows.append([_camel_to_title(key), f"{_fmt_num(val.get('value', '-'))} {val.get('unit', '')}"])
                else:
                    mb_rows.append([_camel_to_title(key), str(val)])
            y = _draw_table(c, ["Parameter", "Value"], mb_rows, left_margin, y,
                            [250, 262], header_bg="#2563EB", page_height=page_height, page_width=page_width)
            y -= 10

        stages = mb_results.get("stages") or []
        equipment = mb_results.get("equipment") or []
        notes = []
        if isinstance(stages, list) and len(stages) > 0:
            notes.append(f"Treatment stages: {len(stages)}")
        if isinstance(equipment, list) and len(equipment) > 0:
            notes.append(f"Equipment items: {len(equipment)}")
        if notes:
            c.setFont("Helvetica", 9)
            c.setFillColor(HexColor("#666666"))
            c.drawString(left_margin, y, " | ".join(notes))
            y -= 14
        y -= 5

    if capex_results and isinstance(capex_results, dict):
        y = _add_section_header(c, "Capital Cost Summary", y, left_margin, content_width, page_height)
        cap_summary = capex_results.get("summary") or {}
        cap_rows = []
        if isinstance(cap_summary, dict):
            eq = cap_summary.get("totalEquipmentCost", 0)
            is_det = "subtotalDirectCosts" in cap_summary
            td = cap_summary.get("subtotalDirectCosts", 0) if is_det else cap_summary.get("totalInstalledCost", 0)
            inst = td - eq
            ec = cap_summary.get("engineeringCost", 0)
            cont = cap_summary.get("contingency", cap_summary.get("totalContingency", 0))
            tc = cap_summary.get("totalProjectCost", 0)
            oi = tc - td - ec - cont
            cap_rows.append(["Equipment", _fmt_currency(eq)])
            cap_rows.append(["Installation", _fmt_currency(inst)])
            cap_rows.append(["Total Direct Costs", _fmt_currency(td)])
            cap_rows.append([f"Engineering ({cap_summary.get('engineeringPct', 7)}%)", _fmt_currency(ec)])
            cap_rows.append(["Other Indirect Costs", _fmt_currency(max(oi, 0))])
            cap_rows.append(["Contingency (7.5%)", _fmt_currency(cont)])
            cap_rows.append(["Total Capital Costs", _fmt_currency(tc)])
            cpu = cap_summary.get("costPerUnit")
            if isinstance(cpu, dict):
                cap_rows.append([f"Cost per Unit ({cpu.get('basis', '')})",
                                 f"{_fmt_currency(cpu.get('value', 0))} / {cpu.get('unit', '')}"])
        if cap_rows:
            y = _draw_table(c, ["Item", "Amount"], cap_rows, left_margin, y,
                            [300, 212], header_bg="#2563EB", page_height=page_height, page_width=page_width)
            y -= 10

        methodology = capex_results.get("methodology")
        if methodology:
            if y < 80:
                c.showPage()
                y = page_height - 50
            c.setFont("Helvetica-Oblique", 8)
            c.setFillColor(HexColor("#666666"))
            c.drawString(left_margin, y, f"Methodology: {_sanitize(str(methodology)[:120])}")
            y -= 14
        y -= 5

    if opex_results and isinstance(opex_results, dict):
        y = _add_section_header(c, "Operating Cost Summary", y, left_margin, content_width, page_height)
        opex_summary = opex_results.get("summary") or opex_results
        opex_rows = []

        total_opex = opex_summary.get("totalAnnualOpex") or opex_summary.get("totalOpex")
        if total_opex is not None:
            opex_rows.append(["Total Annual OpEx", _fmt_currency(total_opex)])

        categories = opex_summary.get("categories") or opex_summary.get("breakdown") or {}
        if isinstance(categories, dict):
            for cat_key, cat_val in categories.items():
                if isinstance(cat_val, dict):
                    opex_rows.append([_camel_to_title(cat_key), _fmt_currency(cat_val.get("annual", cat_val.get("total", 0)))])
                else:
                    opex_rows.append([_camel_to_title(cat_key), _fmt_currency(cat_val)])
        elif isinstance(categories, list):
            for cat in categories:
                if isinstance(cat, dict):
                    opex_rows.append([_sanitize(cat.get("name", cat.get("category", ""))),
                                     _fmt_currency(cat.get("annual", cat.get("total", cat.get("cost", 0))))])

        labor = opex_summary.get("laborCost") or opex_summary.get("labor")
        if labor is not None and not isinstance(categories, (dict, list)):
            opex_rows.append(["Labor", _fmt_currency(labor)])
        energy = opex_summary.get("energyCost") or opex_summary.get("energy")
        if energy is not None and not isinstance(categories, (dict, list)):
            opex_rows.append(["Energy", _fmt_currency(energy)])
        maintenance = opex_summary.get("maintenanceCost") or opex_summary.get("maintenance")
        if maintenance is not None and not isinstance(categories, (dict, list)):
            opex_rows.append(["Maintenance", _fmt_currency(maintenance)])

        net_opex = opex_summary.get("netAnnualOpex") or opex_summary.get("netOpex")
        if net_opex is not None:
            opex_rows.append(["Net Annual OpEx", _fmt_currency(net_opex)])

        if opex_rows:
            y = _draw_table(c, ["Category", "Annual Cost"], opex_rows, left_margin, y,
                            [300, 212], header_bg="#2563EB", page_height=page_height, page_width=page_width)
            y -= 10
        y -= 5

    if financial_results and isinstance(financial_results, dict):
        y = _add_section_header(c, "Financial Model Summary", y, left_margin, content_width, page_height)

        fin_rows = []
        irr = financial_results.get("irr") or financial_results.get("IRR")
        if irr is not None:
            fin_rows.append(["IRR", f"{_fmt_num(irr, 1)}%"])
        npv = financial_results.get("npv10") or financial_results.get("npv") or financial_results.get("NPV")
        if npv is not None:
            fin_rows.append(["NPV @10%", _fmt_currency(npv)])
        moic = financial_results.get("moic") or financial_results.get("MOIC")
        if moic is not None:
            fin_rows.append(["MOIC", f"{_fmt_num(moic, 2)}x"])
        payback = financial_results.get("paybackYears") or financial_results.get("payback")
        if payback is not None:
            fin_rows.append(["Payback Years", _fmt_num(payback, 1)])

        revenue_market = financial_results.get("revenueMarket") or financial_results.get("market")
        if revenue_market is not None:
            fin_rows.append(["Revenue Market", _sanitize(str(revenue_market))])

        biogas_scfm = financial_results.get("biogasFlowSCFM") or financial_results.get("biogasScfm")
        if biogas_scfm is not None:
            fin_rows.append(["Biogas Flow", f"{_fmt_num(biogas_scfm)} SCFM"])
        rng_prod = financial_results.get("rngProductionMMBTU") or financial_results.get("rngMmbtuDay")
        if rng_prod is not None:
            fin_rows.append(["RNG Production", f"{_fmt_num(rng_prod)} MMBTU/day"])

        credits_45z = financial_results.get("credits45Z") or financial_results.get("enable45Z")
        if credits_45z:
            if isinstance(credits_45z, dict):
                enabled = credits_45z.get("enabled", False)
                annual = credits_45z.get("annualCredit") or credits_45z.get("annual")
                label = "45Z Credits"
                val = _fmt_currency(annual) if annual else ("Enabled" if enabled else "Disabled")
                fin_rows.append([label, val])
            elif isinstance(credits_45z, bool):
                fin_rows.append(["45Z Credits", "Enabled" if credits_45z else "Disabled"])
            else:
                fin_rows.append(["45Z Credits", _sanitize(str(credits_45z))])

        if fin_rows:
            y = _draw_table(c, ["Metric", "Value"], fin_rows, left_margin, y,
                            [300, 212], header_bg="#2563EB", page_height=page_height, page_width=page_width)
            y -= 10

        pro_forma = financial_results.get("proForma") or financial_results.get("proforma") or []
        if isinstance(pro_forma, list) and len(pro_forma) > 0:
            if y < 100:
                c.showPage()
                y = page_height - 50
            c.setFont("Helvetica-Bold", 10)
            c.setFillColor(HexColor("#333333"))
            c.drawString(left_margin, y, "Pro Forma Highlights:")
            y -= 14

            pf_rows = []
            year1 = pro_forma[0] if len(pro_forma) > 0 else None
            year10 = pro_forma[9] if len(pro_forma) > 9 else (pro_forma[-1] if len(pro_forma) > 1 else None)

            if year1 and isinstance(year1, dict):
                yr_label = year1.get("year", 1)
                rev = year1.get("revenue") or year1.get("totalRevenue")
                ebitda = year1.get("ebitda") or year1.get("EBITDA")
                if rev is not None:
                    pf_rows.append([f"Year {yr_label} Revenue", _fmt_currency(rev)])
                if ebitda is not None:
                    pf_rows.append([f"Year {yr_label} EBITDA", _fmt_currency(ebitda)])

            if year10 and isinstance(year10, dict) and year10 is not year1:
                yr_label = year10.get("year", 10)
                rev = year10.get("revenue") or year10.get("totalRevenue")
                ebitda = year10.get("ebitda") or year10.get("EBITDA")
                if rev is not None:
                    pf_rows.append([f"Year {yr_label} Revenue", _fmt_currency(rev)])
                if ebitda is not None:
                    pf_rows.append([f"Year {yr_label} EBITDA", _fmt_currency(ebitda)])

            if pf_rows:
                y = _draw_table(c, ["Period", "Amount"], pf_rows, left_margin, y,
                                [300, 212], header_bg="#2563EB", page_height=page_height, page_width=page_width)

    c.save()
    buf.seek(0)
    return buf.read()
