import json
import re
import os
import uuid
import io
import logging
import tempfile
from datetime import datetime
from typing import Optional, Any

from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Response
from pydantic import BaseModel

from services.storage import storage
from api.validation import (
    validate_and_sanitize_output_specs,
    validate_feedstocks_for_type_a,
    apply_ts_tss_guardrail,
    deduplicate_parameters,
    validate_section_assignment,
)
from services.llm import (
    llm_complete,
    get_available_providers,
    is_provider_available,
    PROVIDER_LABELS,
    LLMProvider,
)
from knowledge_base.feedstock_library import (
    enrich_feedstock_specs,
    match_feedstock_type,
    feedstock_group_labels,
    feedstock_group_order,
)
from knowledge_base.output_criteria_library import (
    enrich_output_specs,
    match_output_type,
    output_group_labels,
    output_group_order,
)
from knowledge_base.default_prompts import DEFAULT_PROMPTS, PROMPT_KEYS, PromptKey
from models.schemas import (
    ProjectCreate,
    ScenarioCreate,
    TextEntryCreate,
    FeedstockEntry,
    PromptTemplateUpdate,
)

logger = logging.getLogger(__name__)

api_router = APIRouter(prefix="/api")

VALID_MODELS = [
    "databricks-gpt-5-2-codex",
    "databricks-claude-opus-4-6",
    "databricks-claude-sonnet-4-5",
]

UPLOAD_DIR = "uploads"
MAX_FILE_SIZE = 50 * 1024 * 1024


# ---------------------------------------------------------------------------
# Utility / Helper Functions
# ---------------------------------------------------------------------------


def format_numeric_value(val: str) -> str:
    if not val:
        return val

    def _replace(match: re.Match) -> str:
        num_str = match.group(0).replace(",", "")
        try:
            num = float(num_str)
            if num == int(num):
                return f"{int(num):,}"
            return f"{num:,.2f}".rstrip("0").rstrip(".")
        except ValueError:
            return match.group(0)

    return re.sub(
        r"(?<![.\d])\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d{4,}(?:\.\d+)?",
        _replace,
        val,
    )


def sanitize_pdf_text(text: str) -> str:
    if not text:
        return text
    replacements = {
        "\u2264": "<=", "\u2265": ">=", "\u226a": "<<", "\u226b": ">>",
        "\u2080": "0", "\u2081": "1", "\u2082": "2", "\u2083": "3",
        "\u2084": "4", "\u2085": "5", "\u2086": "6", "\u2087": "7",
        "\u2088": "8", "\u2089": "9",
        "\u2070": "0", "\u00b9": "1", "\u00b2": "2", "\u00b3": "3",
        "\u2074": "4", "\u2075": "5", "\u2076": "6", "\u2077": "7",
        "\u2078": "8", "\u2079": "9",
        "\u2212": "-", "\u2013": "-", "\u2014": "-",
        "\u2032": "'", "\u2033": '"',
        "\u2026": "...", "\u2022": "*", "\u00b7": ".",
        "\u00d7": "x", "\u00f7": "/", "\u00b1": "+/-",
        "\u00b5": "u",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def post_process_pdf_text(raw_text: str) -> str:
    lines = raw_text.split("\n")
    processed: list[str] = []
    in_table = False
    table_headers: list[str] = []
    tables_found = 0

    table_header_patterns = [
        re.compile(
            r"^(Parameter|Element|Alternative|Item|Component|Pollutant|Constituent|"
            r"Category|Description|Metric|Variable|Criteria|Criterion)\s+",
            re.IGNORECASE,
        ),
    ]
    table_header_keywords = [
        "Units", "Limit", "Current", "Loading", "Low", "High", "Base",
        "CAPEX", "OPEX", "Alt.", "Value", "Result", "Average", "Maximum",
        "Minimum", "Target", "Actual", "Standard", "Concentration", "Flow",
        "Cost", "Price", "Total", "Projected", "Annual",
    ]

    for line in lines:
        stripped = line.strip()
        if not stripped:
            if in_table:
                in_table = False
                table_headers = []
                processed.append("")
            continue

        matches_header = any(p.search(stripped) for p in table_header_patterns)
        has_keywords = any(kw in stripped for kw in table_header_keywords)
        cells = re.split(r"\s{2,}", stripped)
        has_multi_cols = len(cells) >= 3

        if (matches_header and has_keywords) or (has_multi_cols and has_keywords and not in_table):
            if len(cells) >= 3:
                in_table = True
                table_headers = cells
                tables_found += 1
                processed.append("")
                processed.append("| " + " | ".join(table_headers) + " |")
                processed.append("| " + " | ".join("---" for _ in table_headers) + " |")
                continue

        if in_table and stripped and not stripped.startswith("Table") and not re.match(r"^\d+\s+[A-Z]", stripped):
            if 2 <= len(cells) <= len(table_headers) + 2:
                processed.append("| " + " | ".join(cells) + " |")
                continue
            else:
                in_table = False
                table_headers = []

        processed.append(stripped)

    if tables_found > 0:
        logger.info("PDF post-processing: converted %d table(s) to structured format", tables_found)

    result = "\n".join(processed)
    result = re.sub(r"\n{3,}", "\n\n", result).strip()
    return result


async def extract_text_from_file(file_bytes: bytes, mime_type: str, original_name: str) -> Optional[str]:
    try:
        ext = os.path.splitext(original_name)[1].lower()

        if mime_type == "application/pdf" or ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(file_bytes))
            text_parts: list[str] = []
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
            raw_text = "\n".join(text_parts).strip()
            if not raw_text:
                return None
            return post_process_pdf_text(raw_text)

        if mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or ext == ".docx":
            from docx import Document as DocxDocument
            doc = DocxDocument(io.BytesIO(file_bytes))
            text = "\n".join(p.text for p in doc.paragraphs)
            return text.strip() or None

        if mime_type in (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ) or ext in (".xlsx", ".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            text_parts_xl: list[str] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_text: list[str] = []
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(c) if c is not None else "" for c in row]
                    rows_text.append(",".join(row_vals))
                csv_text = "\n".join(rows_text).strip()
                if csv_text:
                    text_parts_xl.append(f"[Sheet: {sheet_name}]\n{csv_text}")
            wb.close()
            return "\n\n".join(text_parts_xl).strip() or None

        if mime_type in ("text/plain", "text/csv") or ext in (".txt", ".csv", ".md"):
            text = file_bytes.decode("utf-8", errors="replace")
            return text.strip() or None

        if ext == ".doc":
            return None

        return None
    except Exception as e:
        logger.error("Error extracting text from file %r (%s): %s", original_name, mime_type, str(e))
        return None


def categorize_input(content: str) -> Optional[str]:
    lower = content.lower()

    if any(kw in lower for kw in ["ton", "gallon", "feedstock", "waste", "manure", "organic", "ts%", "vs/ts", "bod", "cod"]):
        return "feedstock"
    if any(kw in lower for kw in ["rng", "biogas", "discharge", "land appl", "output", "produce", "electricity", "compost"]):
        return "output_requirements"
    if any(kw in lower for kw in ["washington", "location", "site", "city", "county", "state"]):
        return "location"
    if any(kw in lower for kw in ["must", "require", "constraint", "assumption", "limit", "using"]):
        return "constraints"
    return None


def extract_parameters_from_text(entries: list[dict]) -> list[dict]:
    params: list[dict] = []
    content = " ".join(e["content"] for e in entries)
    lower_content = content.lower()

    found_feedstock = False
    found_volume = False

    pattern1 = re.compile(
        r"(\d[\d,]*)\s*(tons?|gallons?|lbs?|pounds?)\s*(?:per\s*year|/year|annually)?\s*(?:of\s+)([a-zA-Z\s]+?)(?:\s+from|\s+waste|\s+material|[.,]|$)",
        re.IGNORECASE,
    )
    for m in pattern1.finditer(content):
        if m.group(1) and m.group(3):
            volume = m.group(1).replace(",", "")
            ftype = m.group(3).strip()
            if volume.isdigit() and len(ftype) > 2 and not found_feedstock:
                params.append({"category": "feedstock", "name": "Feedstock Type", "value": ftype.capitalize(), "source": "user_input", "confidence": "high"})
                found_feedstock = True
            if volume.isdigit() and not found_volume:
                params.append({"category": "feedstock", "name": "Volume/Capacity", "value": volume, "unit": f"{m.group(2)}/year", "source": "user_input", "confidence": "high"})
                found_volume = True

    if not found_feedstock:
        waste_types = [
            (re.compile(r"potato\s*(?:waste|processing|peels?|culls?)", re.I), "Potato Waste"),
            (re.compile(r"dairy\s*(?:manure|waste)", re.I), "Dairy Manure"),
            (re.compile(r"food\s*(?:waste|processing|scraps?)", re.I), "Food Waste"),
            (re.compile(r"organic\s*(?:waste|material)", re.I), "Organic Waste"),
            (re.compile(r"agricultural\s*(?:waste|residue)", re.I), "Agricultural Waste"),
            (re.compile(r"manure", re.I), "Manure"),
        ]
        for pat, val in waste_types:
            if pat.search(content):
                params.append({"category": "feedstock", "name": "Feedstock Type", "value": val, "source": "user_input", "confidence": "high"})
                found_feedstock = True
                break

    if not found_volume:
        vol_pat = re.compile(r"(\d[\d,]*)\s*(tons?|gallons?|lbs?|pounds?)\s*(?:per\s*year|/year|annually)?", re.I)
        for m in vol_pat.finditer(content):
            if m.group(1):
                volume = m.group(1).replace(",", "")
                if volume.isdigit() and int(volume) > 100:
                    params.append({"category": "feedstock", "name": "Volume/Capacity", "value": volume, "unit": f"{m.group(2)}/year", "source": "user_input", "confidence": "high"})
                    found_volume = True
                    break

    tech_patterns = [
        (re.compile(r"(\d+(?:\.\d+)?)\s*%?\s*(?:total\s*)?(?:ts|total\s*solids)", re.I), "Total Solids (TS)", "%"),
        (re.compile(r"vs/ts\s*(?:ratio\s*)?(?:of\s*)?(\d+(?:\.\d+)?)", re.I), "VS/TS Ratio", ""),
        (re.compile(r"(\d+(?:\.\d+)?)\s*vs/ts", re.I), "VS/TS Ratio", ""),
        (re.compile(r"bod\s*(?:of\s*)?(\d+(?:,\d+)?)\s*(mg/l|ppm)?", re.I), "BOD", "mg/L"),
        (re.compile(r"cod\s*(?:of\s*)?(\d+(?:,\d+)?)\s*(mg/l|ppm)?", re.I), "COD", "mg/L"),
        (re.compile(r"c:?n\s*(?:ratio\s*)?(?:of\s*)?(\d+):?(\d+)?", re.I), "C:N Ratio", ""),
    ]
    for pat, name, unit in tech_patterns:
        for m in pat.finditer(content):
            if m.group(1):
                value = m.group(1).replace(",", "")
                try:
                    full_val = f"{m.group(1)}:{m.group(2)}" if m.lastindex and m.lastindex >= 2 and m.group(2) else value
                except IndexError:
                    full_val = value
                p: dict[str, Any] = {"category": "feedstock", "name": name, "value": full_val, "source": "user_input", "confidence": "high"}
                if unit:
                    p["unit"] = unit
                params.append(p)

    loc_patterns = [
        re.compile(r"(?:in|at|near|located\s+in)\s+([A-Z][a-zA-Z\s]+,\s*[A-Z]{2}|[A-Z][a-zA-Z\s]+,\s*[A-Z][a-z]+)"),
        re.compile(r"([A-Z][a-zA-Z]+,\s*(?:Washington|Oregon|California|Idaho|Montana))"),
    ]
    for pat in loc_patterns:
        m = pat.search(content)
        if m and m.group(1):
            params.append({"category": "location", "name": "Project Location", "value": m.group(1).strip(), "source": "user_input", "confidence": "high"})
            break

    if "rng" in lower_content:
        params.append({"category": "output_requirements", "name": "Primary Output", "value": "Renewable Natural Gas (RNG)", "source": "user_input", "confidence": "high"})
    if "land appl" in lower_content:
        params.append({"category": "output_requirements", "name": "Solids Handling", "value": "Land Application", "source": "user_input", "confidence": "high"})
    if "discharge" in lower_content and "wwtp" in lower_content:
        params.append({"category": "output_requirements", "name": "Liquid Handling", "value": "Discharge to Municipal WWTP", "source": "user_input", "confidence": "high"})

    constraint_patterns = [
        re.compile(r"must\s+use\s+([^,.]+)", re.I),
        re.compile(r"using\s+([^,.]+)\s+(?:equipment|technology|digester)", re.I),
        re.compile(r"must\s+meet\s+([^,.]+)", re.I),
    ]
    for pat in constraint_patterns:
        for m in pat.finditer(content):
            if m.group(1):
                params.append({"category": "constraints", "name": "Constraint", "value": m.group(1).strip(), "source": "user_input", "confidence": "high"})

    has_feedstock = any(p["category"] == "feedstock" for p in params)
    has_technical = any(p["name"] in ("Total Solids (TS)", "VS/TS Ratio", "BOD") for p in params)

    if has_feedstock and not has_technical:
        if "potato" in lower_content:
            params.extend([
                {"category": "feedstock", "name": "Total Solids (TS)", "value": "12-18", "unit": "%", "source": "predicted", "confidence": "medium"},
                {"category": "feedstock", "name": "VS/TS Ratio", "value": "0.85-0.92", "unit": "", "source": "predicted", "confidence": "medium"},
                {"category": "feedstock", "name": "C:N Ratio", "value": "20-30", "unit": "", "source": "predicted", "confidence": "low"},
            ])
        elif "dairy" in lower_content or "manure" in lower_content:
            params.extend([
                {"category": "feedstock", "name": "Total Solids (TS)", "value": "8-12", "unit": "%", "source": "predicted", "confidence": "medium"},
                {"category": "feedstock", "name": "VS/TS Ratio", "value": "0.75-0.85", "unit": "", "source": "predicted", "confidence": "medium"},
                {"category": "feedstock", "name": "C:N Ratio", "value": "15-25", "unit": "", "source": "predicted", "confidence": "low"},
            ])

    seen: set[tuple[str, str]] = set()
    unique: list[dict] = []
    for p in params:
        key = (p["name"], p["category"])
        if key not in seen:
            seen.add(key)
            unique.append(p)
    return unique


async def extract_parameters_with_ai(
    entries: list[dict],
    model: str = "databricks-gpt-5-2-codex",
    clarifying_qa: Optional[list[dict]] = None,
    prompt_key: str = "extraction",
) -> list[dict]:
    content = "\n\n".join(e["content"] for e in entries)

    if clarifying_qa:
        qa_parts = []
        for i, qa in enumerate(clarifying_qa):
            if qa.get("answer") and qa["answer"].strip():
                qa_parts.append(f"Q{i+1}: {qa['question']}\nA{i+1}: {qa['answer']}")
        if qa_parts:
            content += "\n\n--- ADDITIONAL CLARIFYING INFORMATION ---\nThe following answers were provided to clarifying questions about this project:\n\n" + "\n\n".join(qa_parts)

    if not content.strip():
        logger.info("AI extraction: No content to extract from")
        return []

    if not is_provider_available(model):
        fallback = get_available_providers()
        if not fallback:
            logger.info("AI extraction: No LLM provider available, using pattern matching")
            return extract_parameters_from_text(entries)
        logger.info("AI extraction: %s not available, falling back to %s", model, fallback[0])
        model = fallback[0]

    logger.info("AI extraction: Starting extraction with %s for content length: %d (prompt: %s)", PROVIDER_LABELS.get(model, model), len(content), prompt_key)

    system_prompt = await get_prompt_template(prompt_key)

    try:
        response = llm_complete(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Carefully analyze the following project description and extract ALL parameters. Be thorough - capture every detail mentioned or clearly implied:\n\n{content}"},
            ],
            max_tokens=16384,
            json_mode=True,
        )

        raw_response = response.get("content", "") or "{}"
        logger.info("AI extraction: Received response from %s, length: %d", PROVIDER_LABELS.get(model, model), len(raw_response))
        logger.info("AI extraction: Token usage - prompt: %s completion: %s", response.get("prompt_tokens"), response.get("completion_tokens"))

        result = json.loads(raw_response)

        if not result.get("parameters") or not isinstance(result["parameters"], list):
            logger.warning("AI returned invalid format, falling back to pattern matching. Raw response: %s", raw_response[:500])
            return extract_parameters_from_text(entries)

        logger.info("AI extraction: Successfully extracted %d parameters", len(result["parameters"]))

        extracted = []
        for p in result["parameters"]:
            name = p.get("name") or p.get("parameter_name") or p.get("parameter") or p.get("label") or p.get("field")
            category = p.get("category")
            value = str(p.get("value", ""))
            if name and category and value is not None:
                item: dict[str, Any] = {
                    "category": category,
                    "name": name,
                    "value": value,
                    "source": "ai_extraction",
                    "confidence": p.get("confidence", "medium"),
                }
                unit = p.get("unit") or p.get("units")
                if unit:
                    item["unit"] = unit
                extracted.append(item)
        return extracted
    except Exception as e:
        logger.error("AI extraction failed, falling back to pattern matching. Error: %s", str(e))
        return extract_parameters_from_text(entries)


async def get_prompt_template(key: str) -> str:
    db_template = storage.get_prompt_template(key)
    if db_template:
        return db_template["template"]
    return DEFAULT_PROMPTS[key]["template"]


def classify_feedstock_param(name: str) -> dict:
    numbered = re.match(r"^Feedstock\s+(\d+)\s+(.+)$", name, re.IGNORECASE)
    if numbered:
        return {"index": int(numbered.group(1)), "clean_name": numbered.group(2).strip()}
    lower = name.lower()
    if "primary" in lower or "feedstock type" in lower:
        clean = re.sub(r"primary\s*", "", name, flags=re.I)
        clean = re.sub(r"feedstock\s*", "", clean, flags=re.I).strip() or "Type"
        return {"index": 1, "clean_name": clean}
    if "secondary" in lower:
        clean = re.sub(r"secondary\s*", "", name, flags=re.I)
        clean = re.sub(r"feedstock\s*", "", clean, flags=re.I).strip() or "Type"
        return {"index": 2, "clean_name": clean}
    if "tertiary" in lower:
        clean = re.sub(r"tertiary\s*", "", name, flags=re.I)
        clean = re.sub(r"feedstock\s*", "", clean, flags=re.I).strip() or "Type"
        return {"index": 3, "clean_name": clean}
    if "number of" in lower or "feedstock source" in lower:
        return {"index": 0, "clean_name": name}
    return {"index": 1, "clean_name": name}


def map_technical_param_name(raw_name: str) -> Optional[str]:
    n = raw_name.lower().strip()
    volume_keywords = ["annual", "quantity", "daily", "average", "generation", "onsite", "number of", "facility type", "source", "herd"]
    if any(kw in n for kw in volume_keywords):
        return None
    if "vs/ts" in n or "vs:ts" in n or "volatile solids to total solids" in n:
        return "VS/TS"
    if "total solids" in n or n in ("ts%", "ts (%)", "ts"):
        return "Total Solids"
    if "volatile solids" in n or n in ("vs", "vs (% of ts)"):
        return "Volatile Solids"
    if ("c:n" in n or "c/n" in n) or ("carbon" in n and "nitrogen" in n):
        return "C:N Ratio"
    if "moisture" in n:
        return "Moisture Content"
    if "bulk density" in n or n == "density":
        return "Bulk Density"
    if "bmp" in n or "biochemical methane" in n or "methane potential" in n:
        return "BMP"
    if "biodegradable fraction" in n or "biodegradability" in n:
        return "Biodegradable Fraction"
    return None


# ---------------------------------------------------------------------------
# Pydantic request body models
# ---------------------------------------------------------------------------


class PreferredModelUpdate(BaseModel):
    model: str


class PromptUpdate(BaseModel):
    template: str


class ClarifyAnswersBody(BaseModel):
    answers: list[dict]


class ChatMessageBody(BaseModel):
    message: str


class TextEntryBody(BaseModel):
    content: str
    category: Optional[str] = None


class ProjectBody(BaseModel):
    name: str
    description: Optional[str] = None


class ScenarioBody(BaseModel):
    name: str
    status: str = "draft"
    preferredModel: str = "databricks-gpt-5-2-codex"


# ---------------------------------------------------------------------------
# LLM Providers & Prompt Management Routes
# ---------------------------------------------------------------------------


@api_router.get("/llm-providers")
async def get_llm_providers():
    available = get_available_providers()
    return {
        "providers": [{"id": p, "label": PROVIDER_LABELS.get(p, p)} for p in available],
        "default": available[0] if available else "databricks-gpt-5-2-codex",
    }


@api_router.get("/prompts")
async def list_prompts():
    try:
        db_templates = storage.get_prompt_templates()
        db_map = {t["key"]: t for t in db_templates}

        result = []
        for key in PROMPT_KEYS:
            defaults = DEFAULT_PROMPTS[key]
            db_entry = db_map.get(key)
            result.append({
                "key": defaults["key"],
                "name": db_entry["name"] if db_entry else defaults["name"],
                "description": db_entry.get("description") if db_entry else defaults["description"],
                "template": db_entry["template"] if db_entry else defaults["template"],
                "isSystemPrompt": defaults["is_system_prompt"],
                "availableVariables": defaults["available_variables"],
                "isCustomized": db_entry is not None,
                "updatedAt": db_entry.get("updated_at") if db_entry else None,
            })
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/prompts/{key}")
async def get_prompt(key: str):
    try:
        if key not in PROMPT_KEYS:
            raise HTTPException(status_code=404, detail="Unknown prompt key")
        defaults = DEFAULT_PROMPTS[key]
        db_entry = storage.get_prompt_template(key)
        return {
            "key": defaults["key"],
            "name": db_entry["name"] if db_entry else defaults["name"],
            "description": db_entry.get("description") if db_entry else defaults["description"],
            "template": db_entry["template"] if db_entry else defaults["template"],
            "defaultTemplate": defaults["template"],
            "isSystemPrompt": defaults["is_system_prompt"],
            "availableVariables": defaults["available_variables"],
            "isCustomized": db_entry is not None,
            "updatedAt": db_entry.get("updated_at") if db_entry else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.patch("/prompts/{key}")
async def update_prompt(key: str, body: PromptUpdate):
    try:
        if key not in PROMPT_KEYS:
            raise HTTPException(status_code=404, detail="Unknown prompt key")
        if not body.template or not body.template.strip():
            raise HTTPException(status_code=400, detail="Template text is required")
        defaults = DEFAULT_PROMPTS[key]
        saved = storage.upsert_prompt_template(
            key=key,
            name=defaults["name"],
            description=defaults["description"],
            template=body.template.strip(),
            is_system_prompt=defaults["is_system_prompt"],
        )
        return {
            **saved,
            "availableVariables": defaults["available_variables"],
            "isCustomized": True,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/prompts/{key}/reset")
async def reset_prompt(key: str):
    try:
        if key not in PROMPT_KEYS:
            raise HTTPException(status_code=404, detail="Unknown prompt key")
        storage.delete_prompt_template(key)
        defaults = DEFAULT_PROMPTS[key]
        return {
            "key": defaults["key"],
            "name": defaults["name"],
            "description": defaults["description"],
            "template": defaults["template"],
            "isSystemPrompt": defaults["is_system_prompt"],
            "availableVariables": defaults["available_variables"],
            "isCustomized": False,
            "updatedAt": None,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.patch("/scenarios/{scenario_id}/preferred-model")
async def update_preferred_model(scenario_id: str, body: PreferredModelUpdate):
    try:
        if body.model not in VALID_MODELS:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid model. Must be one of: {', '.join(VALID_MODELS)}",
            )
        scenario = storage.get_scenario(scenario_id)
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")
        updated = storage.update_scenario(scenario_id, {"preferred_model": body.model})
        return updated
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error updating preferred model: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to update preferred model")


# ---------------------------------------------------------------------------
# Projects CRUD
# ---------------------------------------------------------------------------


@api_router.get("/projects")
async def list_projects():
    try:
        return storage.get_projects()
    except Exception as e:
        logger.error("Error fetching projects: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to fetch projects")


@api_router.get("/projects/{project_id}")
async def get_project(project_id: str):
    try:
        project = storage.get_project(project_id)
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        return project
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error fetching project: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to fetch project")


@api_router.post("/projects", status_code=201)
async def create_project(body: ProjectBody):
    try:
        project = storage.create_project(name=body.name, description=body.description)
        return project
    except Exception as e:
        logger.error("Error creating project: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to create project")


@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str):
    try:
        storage.delete_project(project_id)
        return {"success": True}
    except Exception as e:
        logger.error("Error deleting project: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to delete project")


# ---------------------------------------------------------------------------
# Scenarios CRUD
# ---------------------------------------------------------------------------


@api_router.get("/projects/{project_id}/scenarios")
async def list_scenarios(project_id: str):
    try:
        return storage.get_scenarios(project_id)
    except Exception as e:
        logger.error("Error fetching scenarios: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to fetch scenarios")


@api_router.get("/scenarios/recent")
async def get_recent_scenarios():
    try:
        return storage.get_recent_scenarios()
    except Exception as e:
        logger.error("Error fetching recent scenarios: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to fetch recent scenarios")


@api_router.get("/scenarios/{scenario_id}")
async def get_scenario(scenario_id: str):
    try:
        scenario = storage.get_scenario(scenario_id)
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")
        return scenario
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error fetching scenario: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to fetch scenario")


@api_router.post("/projects/{project_id}/scenarios", status_code=201)
async def create_scenario(project_id: str, body: ScenarioBody):
    try:
        scenario = storage.create_scenario(
            project_id=project_id,
            name=body.name,
            status=body.status,
            preferred_model=body.preferredModel,
        )
        return scenario
    except Exception as e:
        logger.error("Error creating scenario: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to create scenario")


@api_router.delete("/scenarios/{scenario_id}")
async def delete_scenario(scenario_id: str):
    try:
        storage.delete_scenario(scenario_id)
        return {"success": True}
    except Exception as e:
        logger.error("Error deleting scenario: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to delete scenario")


@api_router.post("/scenarios/{scenario_id}/classify")
async def classify_project_type(scenario_id: str):
    try:
        scenario = storage.get_scenario(scenario_id)
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")

        entries = storage.get_text_entries(scenario_id)
        documents = storage.get_documents(scenario_id)
        doc_entries = [
            {"content": f"[From document: {doc['original_name']}]\n{doc['extracted_text']}", "category": None}
            for doc in documents
            if doc.get("extracted_text") and doc["extracted_text"].strip()
        ]

        all_entries = list(entries) + doc_entries
        content = "\n\n".join(e["content"] for e in all_entries)

        if not content.strip():
            raise HTTPException(status_code=400, detail="No input content to analyze. Add text or upload documents first.")

        model = scenario.get("preferred_model") or "databricks-gpt-5-2-codex"

        if not get_available_providers():
            raise HTTPException(status_code=500, detail="No AI provider is configured.")

        system_prompt = await get_prompt_template("classification")

        response = llm_complete(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Here is the project information submitted so far:\n\n{content}"},
            ],
            max_tokens=2048,
            json_mode=True,
        )

        raw_response = response.get("content", "") or "{}"
        logger.info("Classification: Received response from %s, length: %d", response.get("provider", model), len(raw_response))

        try:
            parsed = json.loads(raw_response)
        except json.JSONDecodeError:
            logger.error("Classification: Failed to parse JSON response: %s", raw_response[:300])
            raise HTTPException(status_code=500, detail="Failed to parse classification response from AI.")

        project_type = parsed.get("projectType")
        if not project_type or project_type not in ["A", "B", "C", "D"]:
            raise HTTPException(status_code=500, detail="AI returned an invalid project type.")

        storage.update_scenario_project_type(scenario_id, project_type, False)

        return {
            "projectType": project_type,
            "projectTypeName": parsed.get("projectTypeName", ""),
            "confidence": parsed.get("confidence", "medium"),
            "reasoning": parsed.get("reasoning", ""),
            "provider": response.get("provider", model),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error classifying project type: %s", str(e))
        raise HTTPException(status_code=500, detail=f"Failed to classify project type: {str(e)}")


class ProjectTypeBody(BaseModel):
    projectType: str
    confirmed: Optional[bool] = True


@api_router.patch("/scenarios/{scenario_id}/project-type")
async def update_project_type(scenario_id: str, body: ProjectTypeBody):
    try:
        if body.projectType not in ["A", "B", "C", "D"]:
            raise HTTPException(status_code=400, detail="Invalid project type. Must be 'A', 'B', 'C', or 'D'.")

        scenario = storage.get_scenario(scenario_id)
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")

        updated = storage.update_scenario_project_type(
            scenario_id, body.projectType, body.confirmed is not False
        )
        return updated
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error updating project type: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to update project type")


# ---------------------------------------------------------------------------
# Text Entries CRUD
# ---------------------------------------------------------------------------


@api_router.get("/scenarios/{scenario_id}/text-entries")
async def list_text_entries(scenario_id: str):
    try:
        return storage.get_text_entries(scenario_id)
    except Exception as e:
        logger.error("Error fetching text entries: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to fetch text entries")


@api_router.post("/scenarios/{scenario_id}/text-entries", status_code=201)
async def create_text_entry(scenario_id: str, body: TextEntryBody):
    try:
        category = categorize_input(body.content)
        entry = storage.create_text_entry(
            scenario_id=scenario_id,
            content=body.content,
            category=category,
        )
        return entry
    except Exception as e:
        logger.error("Error creating text entry: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to create text entry")


@api_router.delete("/text-entries/{entry_id}")
async def delete_text_entry(entry_id: str):
    try:
        storage.delete_text_entry(entry_id)
        return {"success": True}
    except Exception as e:
        logger.error("Error deleting text entry: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to delete text entry")


# ---------------------------------------------------------------------------
# Documents Upload & CRUD
# ---------------------------------------------------------------------------


@api_router.get("/scenarios/{scenario_id}/documents")
async def list_documents(scenario_id: str):
    try:
        return storage.get_documents(scenario_id)
    except Exception as e:
        logger.error("Error fetching documents: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to fetch documents")


@api_router.post("/scenarios/{scenario_id}/documents", status_code=201)
async def upload_document(scenario_id: str, file: UploadFile = File(...)):
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="No file uploaded")

        file_bytes = await file.read()
        if len(file_bytes) > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail="File exceeds 50MB limit")

        os.makedirs(UPLOAD_DIR, exist_ok=True)
        stored_filename = uuid.uuid4().hex
        file_path = os.path.join(UPLOAD_DIR, stored_filename)
        with open(file_path, "wb") as f:
            f.write(file_bytes)

        extracted_text: Optional[str] = None
        try:
            extracted_text = await extract_text_from_file(file_bytes, file.content_type or "", file.filename)
            if extracted_text:
                logger.info("Document text extraction: extracted %d chars from %s", len(extracted_text), file.filename)
            else:
                logger.info("Document text extraction: no text extracted from %s (unsupported format or empty)", file.filename)
        except Exception as extract_err:
            logger.error("Document text extraction failed (non-fatal): %s", str(extract_err))

        doc = storage.create_document(
            scenario_id=scenario_id,
            filename=stored_filename,
            original_name=file.filename,
            mime_type=file.content_type or "application/octet-stream",
            size=str(len(file_bytes)),
            extracted_text=extracted_text,
        )
        return doc
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error uploading document: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to upload document")


@api_router.delete("/documents/{doc_id}")
async def delete_document(doc_id: str):
    try:
        storage.delete_document(doc_id)
        return {"success": True}
    except Exception as e:
        logger.error("Error deleting document: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to delete document")


# ---------------------------------------------------------------------------
# Parameters & Clarifying Questions
# ---------------------------------------------------------------------------


@api_router.get("/scenarios/{scenario_id}/parameters")
async def list_parameters(scenario_id: str):
    try:
        return storage.get_parameters(scenario_id)
    except Exception as e:
        logger.error("Error fetching parameters: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to fetch parameters")


@api_router.post("/scenarios/{scenario_id}/clarify")
async def generate_clarifying_questions(scenario_id: str):
    try:
        entries = storage.get_text_entries(scenario_id)
        documents = storage.get_documents(scenario_id)
        doc_entries = [
            {"content": f"[From document: {doc['original_name']}]\n{doc['extracted_text']}", "category": None}
            for doc in documents
            if doc.get("extracted_text") and doc["extracted_text"].strip()
        ]

        all_entries = list(entries) + doc_entries
        content = "\n\n".join(e["content"] for e in all_entries)

        if not content.strip():
            raise HTTPException(status_code=400, detail="No input content to analyze. Add text or upload documents first.")

        scenario = storage.get_scenario(scenario_id)
        model = (scenario or {}).get("preferred_model") or "databricks-gpt-5-2-codex"

        if not get_available_providers():
            raise HTTPException(status_code=500, detail="No AI provider is configured.")

        system_prompt = await get_prompt_template("clarify")

        response = llm_complete(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Here is the project information submitted so far:\n\n{content}"},
            ],
            max_tokens=2048,
            json_mode=True,
        )

        raw_response = response.get("content", "") or "{}"
        try:
            parsed = json.loads(raw_response)
        except json.JSONDecodeError:
            parsed = {"questions": [
                {"question": "What are the specific feedstock types and their expected daily/annual volumes?"},
                {"question": "What is the intended use for the biogas produced (e.g., RNG pipeline injection, electricity generation, flaring)?"},
                {"question": "How will the liquid effluent from the digester be managed (e.g., discharge to municipal WWTP, land application, on-site treatment)?"},
            ]}

        questions = parsed.get("questions", [])
        storage.update_scenario(scenario_id, {
            "clarifying_questions": questions,
            "clarifying_answers": None,
        })

        return {"questions": questions, "provider": response.get("provider", model)}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error generating clarifying questions: %s", str(e))
        raise HTTPException(status_code=500, detail=f"Failed to generate clarifying questions: {str(e)}")


@api_router.post("/scenarios/{scenario_id}/clarify-answers")
async def save_clarify_answers(scenario_id: str, body: ClarifyAnswersBody):
    try:
        scenario = storage.get_scenario(scenario_id)
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")

        storage.update_scenario(scenario_id, {
            "clarifying_questions": scenario.get("clarifying_questions"),
            "clarifying_answers": body.answers,
        })
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error saving clarifying answers: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to save answers")


# ---------------------------------------------------------------------------
# Extract & UPIF Generation — THE CORE ENDPOINT
# ---------------------------------------------------------------------------


@api_router.post("/scenarios/{scenario_id}/extract")
async def extract_parameters(scenario_id: str):
    try:
        entries = storage.get_text_entries(scenario_id)
        documents = storage.get_documents(scenario_id)
        doc_entries = [
            {"content": f"[From document: {doc['original_name']}]\n{doc['extracted_text']}", "category": None}
            for doc in documents
            if doc.get("extracted_text") and doc["extracted_text"].strip()
        ]

        if doc_entries:
            logger.info("Including %d document(s) with extracted text in parameter extraction", len(doc_entries))

        all_entries = list(entries) + doc_entries

        extract_scenario = storage.get_scenario(scenario_id)
        extract_model = (extract_scenario or {}).get("preferred_model") or "databricks-gpt-5-2-codex"
        clarifying_answers = (extract_scenario or {}).get("clarifying_answers") or None

        type_prompt_map = {
            "A": "extraction_type_a",
            "B": "extraction_type_b",
            "C": "extraction_type_c",
            "D": "extraction_type_d",
        }
        project_type = (extract_scenario or {}).get("project_type")
        project_type_confirmed = (extract_scenario or {}).get("project_type_confirmed")
        extraction_prompt_key = (
            type_prompt_map.get(project_type, "extraction")
            if project_type and project_type_confirmed
            else "extraction"
        )
        logger.info(
            "Extraction: projectType=%s, confirmed=%s, using prompt: %s",
            project_type, project_type_confirmed, extraction_prompt_key,
        )

        extracted_params = await extract_parameters_with_ai(all_entries, extract_model, clarifying_answers, extraction_prompt_key)

        storage.delete_parameters(scenario_id)

        valid_params = [p for p in extracted_params if p.get("name") and p.get("category")]
        if len(valid_params) < len(extracted_params):
            logger.info("AI extraction: Filtered out %d parameters with missing name/category", len(extracted_params) - len(valid_params))

        valid_params = deduplicate_parameters(valid_params)
        logger.info("Validation: After deduplication: %d parameters", len(valid_params))

        section_result = validate_section_assignment(valid_params)
        valid_params = section_result["valid"]
        all_validation_warnings: list[dict] = list(section_result["warnings"])
        all_unmapped_specs: dict[str, dict] = {}
        all_performance_targets: list[dict] = []
        if section_result["unmapped"]:
            for p in section_result["unmapped"]:
                k = f"section_{p.get('name', 'unknown')}".replace(" ", "_").lower()
                all_unmapped_specs[k] = {
                    "displayName": p.get("name", ""),
                    "value": p.get("value", ""),
                    "unit": p.get("unit", ""),
                    "source": "ai_extraction",
                    "confidence": p.get("confidence", "low"),
                    "provenance": "Moved to unmapped due to section/unit mismatch",
                    "group": "unmapped",
                    "sortOrder": 999,
                }

        for param in valid_params:
            storage.create_parameter(
                scenario_id=scenario_id,
                category=param["category"],
                name=param["name"],
                value=param.get("value"),
                unit=param.get("unit"),
                source=param.get("source", "ai_extraction"),
                confidence=param.get("confidence"),
            )

        existing_upif = storage.get_upif(scenario_id)

        feedstock_params = [p for p in extracted_params if p.get("category") == "feedstock"]
        feedstock_groups: dict[int, list[dict]] = {}
        for param in feedstock_params:
            classified = classify_feedstock_param(param["name"])
            idx = classified["index"]
            if idx == 0:
                continue
            if idx not in feedstock_groups:
                feedstock_groups[idx] = []
            feedstock_groups[idx].append({
                "clean_name": classified["clean_name"],
                "value": param.get("value", ""),
                "unit": param.get("unit"),
            })

        feedstock_entries: list[dict] = []
        for idx in sorted(feedstock_groups.keys()):
            group = feedstock_groups[idx]
            type_param = next(
                (p for p in group if p["clean_name"].lower() in ("type", "") or "type" in p["clean_name"].lower() or p["clean_name"].lower() == "feedstock"),
                None,
            )
            volume_param = next(
                (p for p in group if any(kw in p["clean_name"].lower() for kw in ("volume", "quantity", "capacity"))),
                None,
            )

            feedstock_type = type_param["value"] if type_param else f"Unknown Feedstock {idx}"
            user_params: dict[str, dict] = {}
            raw_params: dict[str, dict] = {}

            for p in group:
                if p is type_param or p is volume_param:
                    continue
                mapped = map_technical_param_name(p["clean_name"])
                if mapped:
                    user_params[mapped] = {"value": p["value"], "unit": p.get("unit")}
                raw_params[p["clean_name"]] = {"value": p["value"], "unit": p.get("unit", "")}

            specs = enrich_feedstock_specs(feedstock_type, user_params, project_type)
            logger.info("Enrichment: Feedstock %d %r - %d specs", idx, feedstock_type, len(specs))

            entry: dict[str, Any] = {
                "feedstockType": feedstock_type,
                "feedstockVolume": volume_param["value"] if volume_param else None,
                "feedstockUnit": volume_param.get("unit") if volume_param else None,
                "feedstockParameters": raw_params if raw_params else None,
                "feedstockSpecs": specs if specs else None,
            }
            feedstock_entries.append(entry)

        logger.info("Enrichment: Total feedstock entries: %d", len(feedstock_entries))

        type_a_result = validate_feedstocks_for_type_a(feedstock_entries, valid_params, project_type)
        feedstock_entries = type_a_result["feedstocks"]
        all_validation_warnings.extend(type_a_result["warnings"])
        if type_a_result["missingRequired"]:
            for req in type_a_result["missingRequired"]:
                all_validation_warnings.append({
                    "field": req,
                    "section": "Type A Requirements",
                    "message": f"Missing required parameter: {req}",
                    "severity": "error",
                })

        ts_tss_result = apply_ts_tss_guardrail(feedstock_entries, valid_params)
        feedstock_entries = ts_tss_result["feedstocks"]
        all_validation_warnings.extend(ts_tss_result["warnings"])

        location_params = [p for p in extracted_params if p.get("category") == "location"]
        location = ", ".join(p["value"] for p in location_params) if location_params else ""

        output_params = [p for p in extracted_params if p.get("category") == "output_requirements"]

        constraints = []
        for p in extracted_params:
            if p.get("category") == "constraints":
                if p["name"] != "Constraint":
                    constraints.append(f"{p['name']}: {p.get('value', '')}")
                else:
                    constraints.append(p.get("value", ""))

        output_specs: dict[str, dict] = {}
        user_output_criteria: dict[str, dict] = {}
        for param in output_params:
            user_output_criteria[param["name"]] = {"value": param.get("value", ""), "unit": param.get("unit")}

        for param in output_params:
            output_desc = f"{param['name']} {param.get('value', '')}".lower()
            matched = match_output_type(output_desc)
            if matched and matched["name"] not in output_specs:
                enriched = enrich_output_specs(matched["name"], user_output_criteria, location or None)
                output_specs[matched["name"]] = enriched
                logger.info("Output enrichment: Generated %d criteria for %s", len(enriched), matched["name"])

        all_output_text = " ".join(f"{p['name']} {p.get('value', '')}" for p in output_params).lower()
        all_input_text = " ".join(e["content"] for e in all_entries).lower()
        search_text = f"{all_output_text} {all_input_text}"

        rng_keywords = ["rng", "pipeline", "biomethane", "renewable natural gas", "upgraded biogas", "pipeline injection"]
        digestate_keywords = ["digestate", "land application", "biosolids", "compost", "soil amendment", "land apply"]
        is_type_a = project_type == "A"
        effluent_keywords = (
            ["effluent", "discharge to sewer", "indirect discharge", "potw", "pretreatment", "liquid effluent", "centrate", "filtrate"]
            if is_type_a
            else ["effluent", "wwtp", "discharge", "sewer", "wastewater", "liquid effluent", "centrate", "filtrate", "liquid digestate", "treatment plant"]
        )

        rng_profile = "Renewable Natural Gas (RNG) - Pipeline Injection"
        digestate_profile = "Solid Digestate - Land Application"
        effluent_profile = "Liquid Effluent - Discharge to WWTP"

        if rng_profile not in output_specs and any(k in search_text for k in rng_keywords):
            enriched = enrich_output_specs(rng_profile, user_output_criteria, location or None)
            output_specs[rng_profile] = enriched
            logger.info("Output enrichment (keyword fallback): Generated %d criteria for %s", len(enriched), rng_profile)

        if digestate_profile not in output_specs and any(k in search_text for k in digestate_keywords):
            if is_type_a:
                explicit_biosolids = any(kw in search_text for kw in ["biosolids", "municipal sludge", "class a", "class b", "part 503"])
                if explicit_biosolids:
                    enriched = enrich_output_specs(digestate_profile, user_output_criteria, location or None)
                    output_specs[digestate_profile] = enriched
                    logger.info("Output enrichment (keyword fallback, Type A explicit biosolids): Generated %d criteria for %s", len(enriched), digestate_profile)
                else:
                    logger.info("Output enrichment: Skipping digestate/biosolids profile for Type A (wastewater) project — not explicitly requested")
            else:
                enriched = enrich_output_specs(digestate_profile, user_output_criteria, location or None)
                output_specs[digestate_profile] = enriched
                logger.info("Output enrichment (keyword fallback): Generated %d criteria for %s", len(enriched), digestate_profile)
        if effluent_profile not in output_specs and any(k in search_text for k in effluent_keywords):
            enriched = enrich_output_specs(effluent_profile, user_output_criteria, location or None)
            output_specs[effluent_profile] = enriched
            logger.info("Output enrichment (keyword fallback): Generated %d criteria for %s", len(enriched), effluent_profile)

        logger.info("Output enrichment: Total output profiles enriched: %d", len(output_specs))

        if output_specs:
            os_result = validate_and_sanitize_output_specs(output_specs, project_type)
            output_specs = os_result["sanitized"]
            all_unmapped_specs.update(os_result["unmapped"])
            all_performance_targets.extend(os_result["performanceTargets"])
            all_validation_warnings.extend(os_result["warnings"])
            logger.info("Validation: Output specs sanitized — %d unmapped, %d targets, %d warnings",
                        len(os_result["unmapped"]), len(os_result["performanceTargets"]), len(os_result["warnings"]))

        if rng_profile in output_specs:
            rng_specs = output_specs[rng_profile]
            solids_contaminants = [
                key for key, spec in rng_specs.items()
                if any(term in f"{spec.get('displayName', '')} {spec.get('value', '')} {spec.get('unit', '')}".lower()
                       for term in ["% total solids", "% ts", "dewatered", "land application", "cake",
                                    "mg/kg dry weight", "pathogen", "vector attraction", "part 503"])
            ]
            for key in solids_contaminants:
                logger.info('Validation: Removing cross-stream contaminant "%s" from RNG profile', key)
                del rng_specs[key]

        if effluent_profile in output_specs:
            eff_specs = output_specs[effluent_profile]
            removal_keys = []
            for key, spec in eff_specs.items():
                if spec.get("source") == "user_provided" and "%" in str(spec.get("value", "")) and "mg/L" not in str(spec.get("value", "")):
                    v = str(spec.get("value", "")).lower()
                    if "removal" in v or (v.startswith(">") and "%" in v):
                        logger.info('Validation: Removal efficiency "%s: %s" detected in effluent limits — removing',
                                    spec.get("displayName", ""), spec.get("value", ""))
                        removal_keys.append(key)
            for key in removal_keys:
                del eff_specs[key]

        new_output_requirements = "; ".join(
            f"{p['name']}: {p.get('value', '')}{' ' + p['unit'] if p.get('unit') else ''}"
            for p in output_params
        )

        cf = (existing_upif or {}).get("confirmed_fields") or {}
        old_feedstocks = (existing_upif or {}).get("feedstocks") or []
        old_output_specs = (existing_upif or {}).get("output_specs")

        merged_feedstocks = list(feedstock_entries) if feedstock_entries else []
        if cf.get("feedstocks") and existing_upif:
            for idx_str, confirmed in cf["feedstocks"].items():
                idx = int(idx_str)
                old_fs = old_feedstocks[idx] if idx < len(old_feedstocks) else None
                if not old_fs:
                    continue
                if idx >= len(merged_feedstocks):
                    while len(merged_feedstocks) <= idx:
                        merged_feedstocks.append(None)
                    merged_feedstocks[idx] = old_fs
                    continue
                new_fs = merged_feedstocks[idx]
                if not new_fs:
                    merged_feedstocks[idx] = old_fs
                    continue
                if confirmed.get("feedstockType"):
                    new_fs["feedstockType"] = old_fs.get("feedstockType", new_fs["feedstockType"])
                if confirmed.get("feedstockVolume"):
                    new_fs["feedstockVolume"] = old_fs.get("feedstockVolume", new_fs.get("feedstockVolume"))
                if confirmed.get("feedstockUnit"):
                    new_fs["feedstockUnit"] = old_fs.get("feedstockUnit", new_fs.get("feedstockUnit"))
                if confirmed.get("feedstockSpecs") and old_fs.get("feedstockSpecs") and new_fs.get("feedstockSpecs"):
                    for spec_key, is_locked in confirmed["feedstockSpecs"].items():
                        if is_locked and spec_key in old_fs["feedstockSpecs"]:
                            new_fs["feedstockSpecs"][spec_key] = old_fs["feedstockSpecs"][spec_key]

        merged_output_specs: Optional[dict] = output_specs if output_specs else None
        if cf.get("outputSpecs") and old_output_specs:
            for profile, spec_confirms in cf["outputSpecs"].items():
                if profile not in old_output_specs:
                    continue
                has_any = any(v for v in spec_confirms.values())
                if not has_any:
                    continue
                if merged_output_specs is None:
                    merged_output_specs = {}
                if profile not in merged_output_specs:
                    merged_output_specs[profile] = {}
                for spec_key, is_locked in spec_confirms.items():
                    if is_locked and spec_key in old_output_specs[profile]:
                        merged_output_specs[profile][spec_key] = old_output_specs[profile][spec_key]

        merged_location = existing_upif["location"] if cf.get("location") and existing_upif and existing_upif.get("location") else location
        merged_output_req = existing_upif["output_requirements"] if cf.get("outputRequirements") and existing_upif and existing_upif.get("output_requirements") else new_output_requirements

        merged_constraints = list(constraints)
        if cf.get("constraints") and existing_upif and existing_upif.get("constraints"):
            for idx_str, is_locked in cf["constraints"].items():
                idx = int(idx_str)
                if is_locked and idx < len(existing_upif["constraints"]):
                    while len(merged_constraints) <= idx:
                        merged_constraints.append("")
                    merged_constraints[idx] = existing_upif["constraints"][idx]

        merged_primary = merged_feedstocks[0] if merged_feedstocks else None

        upif_data: dict[str, Any] = {
            "scenario_id": scenario_id,
            "feedstock_type": merged_primary.get("feedstockType") if merged_primary else None,
            "feedstock_volume": merged_primary.get("feedstockVolume") if merged_primary else None,
            "feedstock_unit": merged_primary.get("feedstockUnit") if merged_primary else None,
            "feedstock_parameters": merged_primary.get("feedstockParameters") if merged_primary else None,
            "feedstock_specs": merged_primary.get("feedstockSpecs") if merged_primary else None,
            "feedstocks": merged_feedstocks if merged_feedstocks else None,
            "output_requirements": merged_output_req,
            "output_specs": merged_output_specs,
            "location": merged_location,
            "constraints": merged_constraints,
            "confirmed_fields": cf if cf else None,
            "is_confirmed": False,
            "validation_warnings": all_validation_warnings if all_validation_warnings else None,
            "unmapped_specs": all_unmapped_specs if all_unmapped_specs else None,
            "performance_targets": all_performance_targets if all_performance_targets else None,
        }
        logger.info("Validation summary: %d warnings, %d unmapped, %d performance targets",
                     len(all_validation_warnings), len(all_unmapped_specs), len(all_performance_targets))

        if existing_upif:
            storage.update_upif(scenario_id, upif_data)
        else:
            storage.create_upif(upif_data)

        storage.update_scenario(scenario_id, {"status": "in_review"})

        params = storage.get_parameters(scenario_id)
        return params
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error extracting parameters: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to extract parameters")


# ---------------------------------------------------------------------------
# UPIF CRUD & Confirm
# ---------------------------------------------------------------------------


@api_router.get("/scenarios/{scenario_id}/upif")
async def get_upif(scenario_id: str):
    try:
        upif = storage.get_upif(scenario_id)
        return upif
    except Exception as e:
        logger.error("Error fetching UPIF: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to fetch UPIF")


@api_router.patch("/scenarios/{scenario_id}/upif")
async def update_upif(scenario_id: str, body: dict = {}):
    try:
        upif = storage.update_upif(scenario_id, body)
        if not upif:
            raise HTTPException(status_code=404, detail="UPIF not found")
        return upif
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error updating UPIF: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to update UPIF")


@api_router.post("/scenarios/{scenario_id}/confirm")
async def confirm_scenario(scenario_id: str):
    try:
        storage.confirm_upif(scenario_id)
        storage.update_scenario(scenario_id, {
            "status": "confirmed",
            "confirmed_at": datetime.utcnow(),
        })
        scenario = storage.get_scenario(scenario_id)
        return scenario
    except Exception as e:
        logger.error("Error confirming scenario: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to confirm scenario")


# ---------------------------------------------------------------------------
# UPIF Reviewer Chat
# ---------------------------------------------------------------------------


@api_router.get("/scenarios/{scenario_id}/upif/chat")
async def get_chat_messages(scenario_id: str):
    try:
        return storage.get_chat_messages(scenario_id)
    except Exception as e:
        logger.error("Error fetching chat messages: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to fetch chat messages")


@api_router.post("/scenarios/{scenario_id}/upif/chat")
async def post_chat_message(scenario_id: str, body: ChatMessageBody):
    try:
        message = body.message
        if not message or not message.strip():
            raise HTTPException(status_code=400, detail="Message is required")

        upif = storage.get_upif(scenario_id)
        if not upif:
            raise HTTPException(status_code=404, detail="No UPIF found for this scenario")

        cf = upif.get("confirmed_fields") or {}
        feedstocks = upif.get("feedstocks") or []

        storage.create_chat_message(
            scenario_id=scenario_id,
            role="user",
            content=message.strip(),
        )

        scenario = storage.get_scenario(scenario_id)
        chat_model = (scenario or {}).get("preferred_model") or "databricks-gpt-5-2-codex"

        if not get_available_providers():
            assistant_msg = storage.create_chat_message(
                scenario_id=scenario_id,
                role="assistant",
                content="No AI provider is configured. Please ensure an API key is set up.",
            )
            return assistant_msg

        locked_fields_list: list[str] = []
        if cf.get("location"):
            locked_fields_list.append("location")
        if cf.get("outputRequirements"):
            locked_fields_list.append("outputRequirements")
        if cf.get("constraints"):
            for idx_str, locked in cf["constraints"].items():
                idx = int(idx_str)
                if locked and upif.get("constraints") and idx < len(upif["constraints"]):
                    locked_fields_list.append(f'constraints[{idx_str}]: "{upif["constraints"][idx]}"')
        if cf.get("feedstocks"):
            for idx_str, fs_conf in cf["feedstocks"].items():
                idx = int(idx_str)
                fs = feedstocks[idx] if idx < len(feedstocks) else None
                if not fs:
                    continue
                if fs_conf.get("feedstockType"):
                    locked_fields_list.append(f'feedstocks[{idx_str}].feedstockType: "{fs.get("feedstockType", "")}"')
                if fs_conf.get("feedstockVolume"):
                    locked_fields_list.append(f'feedstocks[{idx_str}].feedstockVolume: "{fs.get("feedstockVolume", "")}"')
                if fs_conf.get("feedstockUnit"):
                    locked_fields_list.append(f'feedstocks[{idx_str}].feedstockUnit: "{fs.get("feedstockUnit", "")}"')
                if fs_conf.get("feedstockSpecs"):
                    for spec_key, locked in fs_conf["feedstockSpecs"].items():
                        if locked:
                            locked_fields_list.append(f"feedstocks[{idx_str}].feedstockSpecs.{spec_key}")
        if cf.get("outputSpecs"):
            for profile, specs in cf["outputSpecs"].items():
                for spec_key, locked in specs.items():
                    if locked:
                        locked_fields_list.append(f'outputSpecs["{profile}"].{spec_key}')

        upif_snapshot: dict[str, Any] = {
            "location": upif.get("location"),
            "outputRequirements": upif.get("output_requirements"),
            "constraints": upif.get("constraints"),
            "feedstocks": [
                {
                    "index": i,
                    "feedstockType": fs.get("feedstockType"),
                    "feedstockVolume": fs.get("feedstockVolume"),
                    "feedstockUnit": fs.get("feedstockUnit"),
                    "feedstockSpecs": {
                        k: {"value": v.get("value"), "unit": v.get("unit")}
                        for k, v in fs["feedstockSpecs"].items()
                    } if fs.get("feedstockSpecs") else None,
                }
                for i, fs in enumerate(feedstocks)
            ] if feedstocks else [],
            "outputSpecs": {
                profile: {
                    k: {"value": v.get("value"), "unit": v.get("unit")}
                    for k, v in specs.items()
                }
                for profile, specs in upif["output_specs"].items()
            } if upif.get("output_specs") else None,
        }

        chat_history = storage.get_chat_messages(scenario_id)
        recent_history = chat_history[-10:]

        reviewer_template = await get_prompt_template("reviewer_chat")
        system_prompt = reviewer_template.replace(
            "{{UPIF_STATE}}", json.dumps(upif_snapshot, indent=2, default=str)
        ).replace(
            "{{LOCKED_FIELDS}}",
            "\n".join(f"- {f}" for f in locked_fields_list) if locked_fields_list else "None - all fields are unlocked",
        )

        messages: list[dict[str, str]] = [{"role": "system", "content": system_prompt}]
        for msg in recent_history:
            if msg["role"] in ("user", "assistant"):
                messages.append({"role": msg["role"], "content": msg["content"]})
        messages.append({"role": "user", "content": message.strip()})

        response = llm_complete(
            model=chat_model,
            messages=messages,
            max_tokens=8192,
            json_mode=True,
        )

        raw_response = response.get("content", "") or "{}"
        try:
            parsed = json.loads(raw_response)
        except json.JSONDecodeError:
            parsed = {"assistantMessage": "I had trouble processing your request. Please try again."}

        assistant_message = parsed.get("assistantMessage", "I've reviewed your feedback.")
        updates = parsed.get("updates") or {}
        model_changed_fields = parsed.get("changedFields") or []

        patch_data: dict[str, Any] = {}
        actual_changes: list[str] = []

        if updates.get("location") is not None and isinstance(updates["location"], str) and not cf.get("location"):
            patch_data["location"] = updates["location"]
            actual_changes.append("location")

        if updates.get("outputRequirements") is not None and isinstance(updates["outputRequirements"], str) and not cf.get("outputRequirements"):
            patch_data["output_requirements"] = updates["outputRequirements"]
            actual_changes.append("outputRequirements")

        if updates.get("constraints") is not None and isinstance(updates["constraints"], list):
            new_constraints = list(updates["constraints"])
            if cf.get("constraints") and upif.get("constraints"):
                for idx_str, is_locked in cf["constraints"].items():
                    idx = int(idx_str)
                    if is_locked and idx < len(upif["constraints"]):
                        while len(new_constraints) <= idx:
                            new_constraints.append("")
                        new_constraints[idx] = upif["constraints"][idx]
            patch_data["constraints"] = new_constraints
            actual_changes.append("constraints")

        if updates.get("feedstocks") is not None and isinstance(updates["feedstocks"], list):
            new_feedstocks = updates["feedstocks"]
            if cf.get("feedstocks"):
                for idx_str, fs_conf in cf["feedstocks"].items():
                    idx = int(idx_str)
                    old_fs = feedstocks[idx] if idx < len(feedstocks) else None
                    new_fs = new_feedstocks[idx] if idx < len(new_feedstocks) else None
                    if not old_fs or not new_fs:
                        continue
                    if fs_conf.get("feedstockType"):
                        new_fs["feedstockType"] = old_fs.get("feedstockType", "")
                    if fs_conf.get("feedstockVolume"):
                        new_fs["feedstockVolume"] = old_fs.get("feedstockVolume")
                    if fs_conf.get("feedstockUnit"):
                        new_fs["feedstockUnit"] = old_fs.get("feedstockUnit")
                    if fs_conf.get("feedstockSpecs") and old_fs.get("feedstockSpecs") and new_fs.get("feedstockSpecs"):
                        for spec_key, locked in fs_conf["feedstockSpecs"].items():
                            if locked and spec_key in old_fs["feedstockSpecs"]:
                                new_fs["feedstockSpecs"][spec_key] = old_fs["feedstockSpecs"][spec_key]
            patch_data["feedstocks"] = new_feedstocks
            primary = new_feedstocks[0] if new_feedstocks else None
            if primary:
                patch_data["feedstock_type"] = primary.get("feedstockType")
                patch_data["feedstock_volume"] = primary.get("feedstockVolume")
                patch_data["feedstock_unit"] = primary.get("feedstockUnit")
                patch_data["feedstock_specs"] = primary.get("feedstockSpecs")
            actual_changes.append("feedstocks")

        if updates.get("outputSpecs") is not None and isinstance(updates["outputSpecs"], dict):
            new_output_specs = updates["outputSpecs"]
            if cf.get("outputSpecs"):
                old_os = upif.get("output_specs") or {}
                for profile, spec_confirms in cf["outputSpecs"].items():
                    for spec_key, locked in spec_confirms.items():
                        if locked and profile in old_os and spec_key in old_os[profile] and profile in new_output_specs:
                            new_output_specs[profile][spec_key] = old_os[profile][spec_key]
            patch_data["output_specs"] = new_output_specs
            actual_changes.append("outputSpecs")

        if patch_data:
            storage.update_upif(scenario_id, patch_data)

        detailed_changes = (
            [f for f in model_changed_fields if any(f.startswith(ac) or f == ac for ac in actual_changes)]
            if model_changed_fields
            else actual_changes
        )

        assistant_msg = storage.create_chat_message(
            scenario_id=scenario_id,
            role="assistant",
            content=assistant_message,
            applied_updates={
                "changedFields": detailed_changes if detailed_changes else actual_changes,
                "summary": f"Updated: {', '.join(actual_changes)}",
            } if actual_changes else None,
        )

        return assistant_msg
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error in UPIF chat: %s", str(e))
        raise HTTPException(status_code=500, detail=f"Failed to process chat message: {str(e)}")


# ---------------------------------------------------------------------------
# PDF Export
# ---------------------------------------------------------------------------


def _draw_table(
    c,
    headers: list[str],
    rows: list[list[str]],
    start_x: float,
    start_y: float,
    col_widths: list[float],
    font_size: int = 8,
    margin: float = 50,
) -> float:
    from reportlab.lib.colors import HexColor

    min_row_height = 18
    cell_padding = 3
    page_height = 792
    table_width = sum(col_widths)
    y = start_y

    def measure_row_height(cells: list[str], bold: bool) -> float:
        from reportlab.pdfbase.pdfmetrics import stringWidth
        max_h = min_row_height
        font_name = "Helvetica-Bold" if bold else "Helvetica"
        for i, cell_text in enumerate(cells):
            cell_w = col_widths[i] - cell_padding * 2
            text = sanitize_pdf_text(cell_text or "")
            text_w = stringWidth(text, font_name, font_size)
            lines = max(1, int(text_w / cell_w) + 1) if cell_w > 0 else 1
            cell_h = lines * (font_size + 2) + cell_padding * 2
            if cell_h > max_h:
                max_h = cell_h
        return max_h

    def draw_row(cells: list[str], bold: bool, bg_color: Optional[str] = None):
        nonlocal y, c
        row_height = measure_row_height(cells, bold)

        if y - row_height < margin + 30:
            c.showPage()
            y = page_height - margin

        if bg_color:
            c.setFillColor(HexColor(bg_color))
            c.rect(start_x, y - row_height, table_width, row_height, fill=1, stroke=0)

        font_name = "Helvetica-Bold" if bold else "Helvetica"
        c.setFont(font_name, font_size)
        c.setFillColor(HexColor("#333333"))

        x = start_x
        for i, cell_text in enumerate(cells):
            text = sanitize_pdf_text(cell_text or "")
            cell_w = col_widths[i] - cell_padding * 2
            if cell_w > 0 and len(text) > 0:
                from reportlab.pdfbase.pdfmetrics import stringWidth
                text_w = stringWidth(text, font_name, font_size)
                if text_w > cell_w:
                    ratio = cell_w / text_w
                    chars = max(1, int(len(text) * ratio) - 3)
                    if i < len(cells) - 1:
                        text = text[:chars] + "..."
                c.drawString(x + cell_padding, y - cell_padding - font_size, text)
            x += col_widths[i]

        c.setStrokeColor(HexColor("#cccccc"))
        c.setLineWidth(0.5)
        c.line(start_x, y - row_height, start_x + table_width, y - row_height)
        y -= row_height

    draw_row(headers, True, "#e8e8e8")
    for idx, row in enumerate(rows):
        draw_row(row, False, "#f5f5f5" if idx % 2 == 1 else None)

    return y


@api_router.get("/scenarios/{scenario_id}/upif/export-pdf")
async def export_pdf(scenario_id: str):
    try:
        from reportlab.lib.pagesizes import LETTER
        from reportlab.pdfgen import canvas
        from reportlab.lib.colors import HexColor

        scenario = storage.get_scenario(scenario_id)
        if not scenario:
            raise HTTPException(status_code=404, detail="Scenario not found")
        project = storage.get_project(scenario["project_id"])
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")
        upif = storage.get_upif(scenario_id)
        if not upif:
            raise HTTPException(status_code=404, detail="No UPIF found for this scenario")

        is_draft = scenario.get("status") != "confirmed"

        feedstocks_data = upif.get("feedstocks") or []
        if not feedstocks_data and upif.get("feedstock_type"):
            feedstocks_data = [{
                "feedstockType": upif["feedstock_type"],
                "feedstockVolume": upif.get("feedstock_volume"),
                "feedstockUnit": upif.get("feedstock_unit"),
                "feedstockSpecs": upif.get("feedstock_specs"),
            }]

        ai_summary = ""
        try:
            feedstock_desc = ", ".join(
                f"{f.get('feedstockType', 'Unknown')}"
                + (f" ({format_numeric_value(f['feedstockVolume'])} {f.get('feedstockUnit', '')})" if f.get("feedstockVolume") else "")
                for f in feedstocks_data
            )

            pdf_template = await get_prompt_template("pdf_summary")
            prompt = (
                pdf_template
                .replace("{{PROJECT_NAME}}", project["name"])
                .replace("{{SCENARIO_NAME}}", scenario["name"])
                .replace("{{FEEDSTOCKS}}", feedstock_desc or "Not specified")
                .replace("{{LOCATION}}", upif.get("location") or "Not specified")
                .replace("{{OUTPUT_REQUIREMENTS}}", upif.get("output_requirements") or "Not specified")
                .replace("{{CONSTRAINTS}}", "; ".join(upif.get("constraints") or []) or "None specified")
            )

            pdf_model = scenario.get("preferred_model") or "databricks-gpt-5-2-codex"
            completion = llm_complete(
                model=pdf_model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=300,
            )
            ai_summary = (completion.get("content") or "").strip()
        except Exception as err:
            logger.error("LLM summary generation failed, using fallback: %s", str(err))

        if not ai_summary:
            parts: list[str] = []
            parts.append(f'This project intake form documents the "{project["name"]}" project (scenario: "{scenario["name"]}").')
            if feedstocks_data:
                desc = ", ".join(
                    f"{f.get('feedstockType', 'Unknown')}"
                    + (f" at {format_numeric_value(f['feedstockVolume'])} {f.get('feedstockUnit', '')}" if f.get("feedstockVolume") else "")
                    for f in feedstocks_data
                )
                parts.append(f"The proposed feedstock(s) include {desc}.")
            if upif.get("location"):
                parts.append(f"The project is located in {upif['location']}.")
            if upif.get("output_requirements"):
                parts.append(f"Output requirements: {upif['output_requirements']}.")
            if upif.get("constraints"):
                parts.append(f"Key constraints: {'; '.join(upif['constraints'])}.")
            ai_summary = " ".join(parts)

        page_width, page_height = LETTER
        content_width = page_width - 100
        left_margin = 50

        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=LETTER)

        def add_watermark():
            if not is_draft:
                return
            c.saveState()
            c.setFont("Helvetica-Bold", 120)
            c.setFillColor(HexColor("#e53e3e"))
            c.setFillAlpha(0.12)
            c.translate(page_width / 2, page_height / 2)
            c.rotate(45)
            c.drawCentredString(0, 0, "DRAFT")
            c.restoreState()

        add_watermark()

        current_y = page_height - 50

        c.setFont("Helvetica-Bold", 18)
        c.setFillColor(HexColor("#222222"))
        c.drawCentredString(page_width / 2, current_y, "UNIFIED PROJECT INTAKE FORM")
        current_y -= 24

        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(HexColor("#444444"))
        c.drawCentredString(page_width / 2, current_y, project["name"])
        current_y -= 18

        c.setFont("Helvetica", 11)
        c.setFillColor(HexColor("#666666"))
        c.drawCentredString(page_width / 2, current_y, f"Scenario: {scenario['name']}")
        current_y -= 16

        pdf_project_type = scenario.get("project_type")
        pdf_type_confirmed = scenario.get("project_type_confirmed")
        pdf_type_labels = {"A": "Wastewater Treatment", "B": "RNG Greenfield", "C": "RNG Bolt-On", "D": "Hybrid"}
        if pdf_project_type and pdf_type_confirmed:
            c.setFont("Helvetica-Bold", 11)
            c.setFillColor(HexColor("#2563eb"))
            c.drawCentredString(page_width / 2, current_y, f"Project Type {pdf_project_type}: {pdf_type_labels.get(pdf_project_type, pdf_project_type)}")
            current_y -= 16

        date_str = ""
        if upif.get("created_at"):
            try:
                dt = upif["created_at"]
                if isinstance(dt, str):
                    dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
                date_str = dt.strftime("%B %d, %Y")
            except Exception:
                date_str = str(upif["created_at"])
        c.setFont("Helvetica", 9)
        c.setFillColor(HexColor("#888888"))
        c.drawRightString(left_margin + content_width, current_y, date_str)
        current_y -= 14

        status_text = "DRAFT" if is_draft else "CONFIRMED"
        status_color = "#d97706" if is_draft else "#16a34a"
        c.setFont("Helvetica-Bold", 9)
        c.setFillColor(HexColor(status_color))
        c.drawRightString(left_margin + content_width, current_y, f"Status: {status_text}")
        current_y -= 14

        c.setStrokeColor(HexColor("#cccccc"))
        c.setLineWidth(1)
        c.line(left_margin, current_y, left_margin + content_width, current_y)
        current_y -= 20

        c.setFont("Helvetica-Bold", 13)
        c.setFillColor(HexColor("#222222"))
        c.drawString(left_margin, current_y, "Project Summary")
        current_y -= 16

        c.setFont("Helvetica", 10)
        c.setFillColor(HexColor("#333333"))
        summary_text = sanitize_pdf_text(format_numeric_value(ai_summary))
        text_obj = c.beginText(left_margin, current_y)
        text_obj.setFont("Helvetica", 10)
        text_obj.setFillColor(HexColor("#333333"))
        text_obj.setLeading(14)
        for line in _wrap_text(summary_text, content_width, "Helvetica", 10):
            text_obj.textLine(line)
        c.drawText(text_obj)
        current_y = text_obj.getY() - 16

        c.setStrokeColor(HexColor("#cccccc"))
        c.setLineWidth(0.5)
        c.line(left_margin, current_y, left_margin + content_width, current_y)
        current_y -= 20

        c.setFont("Helvetica-Bold", 13)
        c.setFillColor(HexColor("#222222"))
        c.drawString(left_margin, current_y, "Feedstock Information")
        current_y -= 16

        if not feedstocks_data:
            c.setFont("Helvetica", 10)
            c.setFillColor(HexColor("#666666"))
            c.drawString(left_margin, current_y, "No feedstock information available.")
            current_y -= 14
        else:
            for feedstock in feedstocks_data:
                if current_y < 100:
                    c.showPage()
                    add_watermark()
                    current_y = page_height - 50

                c.setFont("Helvetica-Bold", 11)
                c.setFillColor(HexColor("#333333"))
                c.drawString(left_margin, current_y, sanitize_pdf_text(feedstock.get("feedstockType", "Unknown")))
                current_y -= 16

                if feedstock.get("feedstockVolume"):
                    c.setFont("Helvetica", 10)
                    c.setFillColor(HexColor("#555555"))
                    c.drawString(
                        left_margin,
                        current_y,
                        sanitize_pdf_text(f"Volume: {format_numeric_value(feedstock['feedstockVolume'])} {feedstock.get('feedstockUnit', '')}"),
                    )
                    current_y -= 14

                if feedstock.get("feedstockSpecs") and len(feedstock["feedstockSpecs"]) > 0:
                    specs = feedstock["feedstockSpecs"]
                    grouped: dict[str, list[tuple[str, Any]]] = {}
                    for key, spec in specs.items():
                        group = spec.get("group", "extended")
                        if group not in grouped:
                            grouped[group] = []
                        grouped[group].append((key, spec))

                    for group in grouped:
                        grouped[group].sort(key=lambda x: x[1].get("sortOrder", 0))

                    col_widths = [130, 100, 60, 70, 152]
                    headers = ["Parameter", "Value", "Unit", "Source", "Notes"]

                    for group_key in feedstock_group_order:
                        items = grouped.get(group_key)
                        if not items:
                            continue

                        if current_y < 100:
                            c.showPage()
                            add_watermark()
                            current_y = page_height - 50

                        c.setFont("Helvetica-Bold", 9)
                        c.setFillColor(HexColor("#555555"))
                        c.drawString(left_margin, current_y, feedstock_group_labels.get(group_key, group_key))
                        current_y -= 12

                        table_rows = [
                            [
                                spec.get("displayName", ""),
                                format_numeric_value(spec.get("value", "")),
                                spec.get("unit", ""),
                                "User" if spec.get("source") == "user_provided" else "Estimated",
                                spec.get("provenance", ""),
                            ]
                            for _, spec in items
                        ]
                        current_y = _draw_table(c, headers, table_rows, left_margin, current_y, col_widths)
                        current_y -= 6

                current_y -= 8

        if current_y < 100:
            c.showPage()
            add_watermark()
            current_y = page_height - 50

        c.setStrokeColor(HexColor("#cccccc"))
        c.setLineWidth(0.5)
        c.line(left_margin, current_y, left_margin + content_width, current_y)
        current_y -= 20

        c.setFont("Helvetica-Bold", 13)
        c.setFillColor(HexColor("#222222"))
        c.drawString(left_margin, current_y, "Output Requirements & Acceptance Criteria")
        current_y -= 16

        if upif.get("output_requirements"):
            c.setFont("Helvetica", 10)
            c.setFillColor(HexColor("#333333"))
            text_obj = c.beginText(left_margin, current_y)
            text_obj.setFont("Helvetica", 10)
            text_obj.setLeading(14)
            for line in _wrap_text(upif["output_requirements"], content_width, "Helvetica", 10):
                text_obj.textLine(line)
            c.drawText(text_obj)
            current_y = text_obj.getY() - 14

        if upif.get("output_specs") and len(upif["output_specs"]) > 0:
            for profile_name, criteria in upif["output_specs"].items():
                if current_y < 100:
                    c.showPage()
                    add_watermark()
                    current_y = page_height - 50

                c.setFont("Helvetica-Bold", 11)
                c.setFillColor(HexColor("#333333"))
                c.drawString(left_margin, current_y, sanitize_pdf_text(profile_name))
                current_y -= 14

                grouped_out: dict[str, list[tuple[str, Any]]] = {}
                for key, spec in criteria.items():
                    group = spec.get("group", "regulatory")
                    if group not in grouped_out:
                        grouped_out[group] = []
                    grouped_out[group].append((key, spec))

                for group in grouped_out:
                    grouped_out[group].sort(key=lambda x: x[1].get("sortOrder", 0))

                col_widths = [110, 80, 55, 75, 55, 137]
                headers = ["Criterion", "Value", "Unit", "Source", "Confidence", "Notes"]

                for group_key in output_group_order:
                    items = grouped_out.get(group_key)
                    if not items:
                        continue

                    if current_y < 100:
                        c.showPage()
                        add_watermark()
                        current_y = page_height - 50

                    c.setFont("Helvetica-Bold", 9)
                    c.setFillColor(HexColor("#555555"))
                    c.drawString(left_margin, current_y, output_group_labels.get(group_key, group_key))
                    current_y -= 12

                    table_rows = [
                        [
                            spec.get("displayName", ""),
                            format_numeric_value(spec.get("value", "")),
                            spec.get("unit", ""),
                            (spec.get("source", "") or "").replace("_", " "),
                            spec.get("confidence", ""),
                            spec.get("provenance", ""),
                        ]
                        for _, spec in items
                    ]
                    current_y = _draw_table(c, headers, table_rows, left_margin, current_y, col_widths)
                    current_y -= 6
                current_y -= 8

        if current_y < 100:
            c.showPage()
            add_watermark()
            current_y = page_height - 50

        c.setStrokeColor(HexColor("#cccccc"))
        c.setLineWidth(0.5)
        c.line(left_margin, current_y, left_margin + content_width, current_y)
        current_y -= 20

        c.setFont("Helvetica-Bold", 13)
        c.setFillColor(HexColor("#222222"))
        c.drawString(left_margin, current_y, "Location")
        current_y -= 14

        c.setFont("Helvetica", 10)
        c.setFillColor(HexColor("#333333"))
        c.drawString(left_margin, current_y, sanitize_pdf_text(upif.get("location") or "Not specified"))
        current_y -= 18

        c.setFont("Helvetica-Bold", 13)
        c.setFillColor(HexColor("#222222"))
        c.drawString(left_margin, current_y, "Constraints")
        current_y -= 14

        if upif.get("constraints") and len(upif["constraints"]) > 0:
            for constraint in upif["constraints"]:
                if current_y < 60:
                    c.showPage()
                    add_watermark()
                    current_y = page_height - 50
                c.setFont("Helvetica", 10)
                c.setFillColor(HexColor("#333333"))
                c.drawString(left_margin, current_y, sanitize_pdf_text(f"  *  {constraint}"))
                current_y -= 14
        else:
            c.setFont("Helvetica", 10)
            c.setFillColor(HexColor("#666666"))
            c.drawString(left_margin, current_y, "No constraints specified.")
            current_y -= 14

        page_count = c.getPageNumber()
        c.save()

        buf.seek(0)
        pdf_bytes = buf.read()

        safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", scenario["name"])
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="UPIF-{safe_name}.pdf"',
                "Content-Length": str(len(pdf_bytes)),
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error exporting UPIF PDF: %s", str(e))
        raise HTTPException(status_code=500, detail="Failed to export PDF")


def _wrap_text(text: str, max_width: float, font_name: str, font_size: int) -> list[str]:
    from reportlab.pdfbase.pdfmetrics import stringWidth
    words = text.split()
    lines: list[str] = []
    current_line = ""
    for word in words:
        test_line = f"{current_line} {word}".strip()
        if stringWidth(test_line, font_name, font_size) <= max_width:
            current_line = test_line
        else:
            if current_line:
                lines.append(current_line)
            current_line = word
    if current_line:
        lines.append(current_line)
    return lines if lines else [""]
